#!/usr/bin/env python3
"""
NGAME Training Flow Manager
Implements Chart 1: Training Flow with CPI-Array Columns Added Iteratively to 17×N Matrix
Manages the complete training workflow for building churn reference array
"""

import os
import json
import numpy as np
import sys
import argparse
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class NGameTrainingFlowManager:
    """
    NGAME Training Flow Manager for Chart 1 implementation.
    Manages the iterative addition of CPI-array columns to 17×N matrix.
    """
    
    def __init__(self):
        self.training_matrix_file = "NGAME_Training_Matrix.xlsx"
        self.target_training_days = 30
        self.current_day = 0
        self.cpi_array_size = 18  # 18 transaction types (incl. Vendors φ18)
        
        # Initialize agents
        self.data_extraction_agent = None
        self.cpi_analysis_agent = None
        self.matrix_management_agent = None
        
    def initialize_agents(self):
        """Initialize all agents for training flow."""
        logger.info("🤖 Initializing Training Flow Agents...")
        
        # Import and initialize agents
        from ngame_data_extraction_agent import NGameDataExtractionAgent
        from ngame_cpi_analysis_agent import NGameCpiAnalysisAgent
        from ngame_matrix_management_agent import NGameMatrixManagementAgent
        
        self.data_extraction_agent = NGameDataExtractionAgent()
        self.cpi_analysis_agent = NGameCpiAnalysisAgent()
        self.matrix_management_agent = NGameMatrixManagementAgent()
        
        logger.info("✅ All Training Flow agents initialized successfully")
    
    def initialize_first_time_matrix(self) -> Dict[str, Any]:
        """
        Initialize the training matrix for first-time setup.
        Creates matrix with only Transaction Type, μ, and σ columns (no daily columns).
        """
        logger.info("🚀 Initializing first-time training matrix")
        print("🚀 First-time Training Matrix Initialization")
        print("=" * 60)
        print("📊 Creating matrix with Transaction Type, μ (Mean CPI), and σ (Std Dev CPI) columns only")
        print("📈 No daily columns will be added initially")
        
        try:
            from ngame_matrix_management_agent import NGameMatrixManagementAgent
            matrix_agent = NGameMatrixManagementAgent()
            
            result = matrix_agent._create_initial_matrix_start_mode()
            
            if result['success']:
                print(f"✅ First-time matrix created successfully!")
                print(f"📊 Matrix dimensions: {result['matrix_dimensions']}")
                print(f"📊 Transaction types: {result['matrix_stats']['transaction_types']}")
                print(f"📊 Training days: {result['matrix_stats']['training_days']}")
                print(f"\n🎯 Matrix ready for training flow")
                print(f"📈 Run without 'start' argument to begin training")
                
                logger.info("✅ First-time matrix initialization completed successfully")
                return result
            else:
                print(f"❌ First-time matrix creation failed: {result.get('error', 'Unknown error')}")
                logger.error(f"❌ First-time matrix initialization failed: {result.get('error', 'Unknown error')}")
                return result
                
        except Exception as e:
            error_msg = f"Error during first-time matrix initialization: {str(e)}"
            print(f"❌ {error_msg}")
            logger.error(f"❌ {error_msg}")
            return {
                'success': False,
                'error': error_msg
            }
        
    def execute_training_day(self, day_number: int) -> Dict[str, Any]:
        """
        Execute training for a single day.
        Implements the training flow from Chart 1.
        """
        logger.info(f"📊 Executing Training Day {day_number}")
        print(f"📊 Training Day {day_number}")
        print("=" * 50)
        
        try:
            # Step 1: Data Extraction (Phase I)
            print(f"📊 Phase I: Data Extraction for Day {day_number}")
            extraction_result = self.data_extraction_agent.extract_daily_data()
            
            if not extraction_result['success']:
                raise Exception(f"Data extraction failed: {extraction_result['error']}")
            
            # Step 2: CPI Analysis (Phase II)
            print(f"🔍 Phase II: CPI Analysis for Day {day_number}")
            cpi_result = self.cpi_analysis_agent.analyze_cpi_coefficients()
            
            if not cpi_result['success']:
                if day_number == 1:
                    logger.info("ℹ️  First day - no CPI analysis (no Yesterday.ttl)")
                    cpi_array          = None
                    weighted_cpi_array = None
                else:
                    raise Exception(f"CPI analysis failed: {cpi_result['error']}")
            else:
                cpi_array          = cpi_result['cpi_array']
                weighted_cpi_array = cpi_result.get('weighted_cpi_array')

            # Step 3: Matrix Update (Training Tail)
            print(f"📈 Training Tail: Matrix Update for Day {day_number}")
            matrix_result = self.matrix_management_agent.update_training_matrix(
                day_number, cpi_array
            )

            # Step 3b: Dollar-weighted baseline update (second sheet)
            if weighted_cpi_array is not None:
                self.matrix_management_agent.update_dollar_weighted_baseline(
                    day_number, weighted_cpi_array
                )
            
            if not matrix_result['success']:
                raise Exception(f"Matrix update failed: {matrix_result['error']}")
            
            # Update current day
            self.current_day = day_number
            
            result = {
                'success': True,
                'day_number': day_number,
                'cpi_array': cpi_array,
                'matrix_updated': True,
                'extraction_result': extraction_result,
                'cpi_result': cpi_result,
                'matrix_result': matrix_result
            }
            
            logger.info(f"✅ Training Day {day_number} completed successfully")
            return result
            
        except Exception as e:
            logger.error(f"❌ Training Day {day_number} failed: {str(e)}")
            return {
                'success': False,
                'day_number': day_number,
                'error': str(e)
            }
    
    def execute_training_period(self, start_day: int = 1, end_day: int = None) -> Dict[str, Any]:
        """
        Execute complete training period.
        Implements the iterative matrix evolution from Chart 1.
        """
        if end_day is None:
            end_day = self.target_training_days
            
        logger.info(f"📊 Starting Training Period: Days {start_day} to {end_day}")
        print(f"📊 NGAME Training Period")
        print("=" * 60)
        print(f"🎯 Building churn reference array")
        print(f"📈 Target training days: {self.target_training_days}")
        print(f"📅 Training period: Day {start_day} to Day {end_day}")
        
        # Initialize agents
        self.initialize_agents()
        
        training_results = {
            'success': True,
            'start_day': start_day,
            'end_day': end_day,
            'days_completed': 0,
            'days_failed': 0,
            'daily_results': [],
            'matrix_evolution': []
        }
        
        for day in range(start_day, end_day + 1):
            print(f"\n📊 Training Day {day}")
            print("-" * 30)
            
            # Execute training day
            day_result = self.execute_training_day(day)
            training_results['daily_results'].append(day_result)
            
            if day_result['success']:
                training_results['days_completed'] += 1
                
                # Track matrix evolution
                matrix_info = {
                    'day': day,
                    'cpi_array': day_result.get('cpi_array'),
                    'matrix_updated': day_result.get('matrix_updated', False)
                }
                training_results['matrix_evolution'].append(matrix_info)
                
                print(f"✅ Day {day} completed successfully")
            else:
                training_results['days_failed'] += 1
                print(f"❌ Day {day} failed: {day_result.get('error', 'Unknown error')}")
        
        # Final training period summary
        self.show_training_period_summary(training_results)
        
        return training_results
    
    def show_training_period_summary(self, training_results: Dict[str, Any]):
        """Show training period summary."""
        print(f"\n📊 Training Period Summary")
        print("=" * 50)
        
        days_completed = training_results['days_completed']
        days_failed = training_results['days_failed']
        total_days = days_completed + days_failed
        
        print(f"📅 Training Period: Day {training_results['start_day']} to Day {training_results['end_day']}")
        print(f"✅ Days Completed: {days_completed}")
        print(f"❌ Days Failed: {days_failed}")
        print(f"📊 Success Rate: {(days_completed / total_days) * 100:.1f}%")
        
        # Matrix evolution summary
        print(f"\n📈 Matrix Evolution Summary")
        print("-" * 30)
        
        for matrix_info in training_results['matrix_evolution']:
            day = matrix_info['day']
            cpi_array = matrix_info['cpi_array']
            matrix_updated = matrix_info['matrix_updated']
            
            if day == 1:
                print(f"Day {day}: Matrix initialized (17×1)")
            elif cpi_array is not None:
                print(f"Day {day}: CPI array added → Matrix (17×{day})")
            else:
                print(f"Day {day}: No CPI array (first day)")
        
        # Training progress
        if days_completed >= self.target_training_days:
            print(f"\n🎉 Training period complete!")
            print(f"✅ Ready for fraud analysis mode")
            print(f"📊 Churn reference array established")
        else:
            remaining_days = self.target_training_days - days_completed
            print(f"\n📅 Remaining training days: {remaining_days}")
            print(f"📊 Continue training to reach {self.target_training_days} days")
    
    def get_training_status(self) -> Dict[str, Any]:
        """Get current training status."""
        status = {
            'current_day': self.current_day,
            'target_days': self.target_training_days,
            'matrix_file': self.training_matrix_file,
            'matrix_exists': os.path.exists(self.training_matrix_file)
        }
        
        if status['matrix_exists']:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(self.training_matrix_file)
                worksheet = workbook.active
                
                # A=type, B=μ, C=σ → Day columns start at col 4
                training_days = worksheet.max_column - 3
                status['training_days'] = training_days
                status['matrix_dimensions'] = f"17×{training_days}"
                status['training_complete'] = training_days >= self.target_training_days
                
            except Exception as e:
                status['error'] = str(e)
        
        return status

def main():
    """Main function for NGAME Training Flow Manager."""
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='NGAME Training Flow Manager')
    parser.add_argument('--start', action='store_true', 
                        help='Initialize first-time matrix with only μ/σ columns (no daily columns)')
    args = parser.parse_args()
    
    print("🚀 NGAME Training Flow Manager")
    print("=" * 60)
    print("📊 Chart 1: Training Flow with Matrix Evolution")
    
    # Initialize training flow manager
    training_manager = NGameTrainingFlowManager()
    
    # Handle first-time initialization
    if args.start:
        result = training_manager.initialize_first_time_matrix()
        return result['success']
    
    # Check current training status
    status = training_manager.get_training_status()
    print(f"\n📊 Current Training Status")
    print(f"Current Day: {status['current_day']}")
    print(f"Target Days: {status['target_days']}")
    
    days_recorded = status.get('training_days', 0)
    target_days   = status['target_days']
    next_day      = days_recorded + 1

    if status['matrix_exists']:
        print(f"Training Days Recorded : {days_recorded} of {target_days}")
        print(f"Matrix Dimensions      : {status.get('matrix_dimensions', 'N/A')}")
    else:
        print("📊 No training matrix found — will create it now")

    # ── Already complete ──────────────────────────────────────────────────────
    if status.get('training_complete', False):
        print("\n✅ Training period complete — ready for fraud analysis")
        return True

    # ── True daily cadence: add exactly ONE day per invocation ────────────────
    print(f"\n📅 Adding Day {next_day} of {target_days} to training matrix...")
    result = training_manager.execute_training_period(next_day, next_day)

    days_remaining = target_days - next_day
    if result.get('success'):
        if days_remaining > 0:
            print(f"\n✅ Day {next_day} recorded successfully.")
            print(f"📅 {days_remaining} day(s) remaining.")
            print(f"   Run this script again tomorrow to add Day {next_day + 1}.")
        else:
            print(f"\n🎉 All {target_days} training days collected!")
            print(f"   Matrix is ready for fraud analysis.")
    else:
        print(f"\n❌ Day {next_day} failed — see errors above.")

    return result.get('success', False)

if __name__ == "__main__":
    main()
