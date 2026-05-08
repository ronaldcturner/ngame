#!/usr/bin/env python3
"""
NGAME Demo Scenario Runner
─────────────────────────────────────────────────────────────────────────────
Runs a controlled fraud-detection demo WITHOUT requiring live TTL snapshots
or real QuickBooks data.

What it does
────────────
1. Reads μ and σ directly from NGAME_Training_Matrix.xlsx (the fabricated or
   real baseline — whichever is present).
2. Builds a "normal day" array  = μ  (every type at baseline).
3. Builds a "fraud day" array   = normal, but three business-meaningful types
   are pushed to μ + SPIKE_SIGMA × σ, guaranteeing HIGH z-score flags.
4. Runs the full churn-comparison → anomaly-id → account-mapping →
   LLM-analysis → management-warning pipeline with the fraud array.
5. Saves all standard output artifacts (NGAME_Fraud_Analysis*.json,
   management_dashboard.json) so the dashboard refreshes correctly.

Injected anomalies — asset misappropriation lecture scenario
─────────────────────────────────────────────────────────────
  φ18 Vendors          z≈5.0 — PRIMARY: fictitious/shell vendor scheme
                                ACFE ranks vendor fraud as the #1 most common
                                asset misappropriation category.
  φ16 ChartOfAccounts  z≈4.0 — CONCEALMENT: chart-of-accounts manipulation
                                New or restructured accounts used to hide the
                                fraudulent cash flows.
  φ14 Contractors      z≈3.5 — SECONDARY: shell contractor / 1099 payments
                                Fictitious or related-party contractor billing,
                                common in non-profits and small businesses.

All other 15 transaction types sit exactly at baseline → LOW z-scores.
"""

import sys
import os

# ── Resolve project root ──────────────────────────────────────────────────────
ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

# ── Configuration ─────────────────────────────────────────────────────────────
MATRIX_FILE  = os.path.join(ROOT, "NGAME_Training_Matrix.xlsx")
# Per-spike sigma multipliers — controls z-score magnitude and rank order.
# Format:  0-based_index: (label, sigma_multiplier)
SPIKE_INDICES = {
    17: ("Vendors",          5.0),   # φ18 — PRIMARY: fictitious/shell vendor scheme   (ranked #1)
    15: ("ChartOfAccounts",  4.0),   # φ16 — CONCEALMENT: account manipulation         (ranked #2)
    13: ("Contractors",      3.5),   # φ14 — SECONDARY: shell contractor / 1099 fraud  (ranked #3)
}

TRANSACTION_TYPES = [
    "Customers", "Recurring_payments", "Invoices", "Payments",
    "Time_Activities", "Bills", "Bill_Payments", "Expenses",
    "Bank_Transactions", "Sales_transactions", "Products",
    "PurchaseOrders", "Recurring_Transactions", "Contractors",
    "Mileage", "ChartOfAccounts", "EmployeePayroll",
    "Vendors",          # φ18 — high-risk fraud hot-spot
]


def load_baseline(matrix_file: str):
    """Return (mu_list, sigma_list) as floats from the training matrix."""
    from openpyxl import load_workbook
    wb = load_workbook(matrix_file)
    ws = wb.active
    mu_list, sigma_list = [], []
    for row in range(2, ws.max_row + 1):
        mu    = ws.cell(row=row, column=2).value
        sigma = ws.cell(row=row, column=3).value
        mu_list.append(float(mu    or 0.0))
        sigma_list.append(float(sigma or 0.0))
    return mu_list, sigma_list


def build_fraud_array(mu, sigma):
    """
    Start from the baseline (μ) then spike the demo indices.
    Each entry in SPIKE_INDICES carries its own sigma multiplier.
    Clamp everything to [0.001, 0.999].
    """
    today = [round(float(m), 6) for m in mu]       # normal baseline
    for idx, (name, multiplier) in SPIKE_INDICES.items():
        s = sigma[idx] if sigma[idx] > 0 else 0.01  # avoid zero-σ
        spiked = mu[idx] + multiplier * s
        today[idx] = round(min(max(spiked, 0.001), 0.999), 6)
        print(f"  💉 Spike injected → φ{idx+1:2d} {name}: "
              f"μ={mu[idx]:.4f}  σ={s:.4f}  "
              f"today={today[idx]:.4f}  "
              f"z≈{multiplier:.1f}")
    return today


def run_pipeline(today_cpi_array):
    """Drive the fraud-analysis pipeline directly (no data-extraction step)."""
    from ngame_churn_comparison_agent    import NGameChurnComparisonAgent
    from ngame_anomaly_identification_agent import NGameAnomalyIdentificationAgent
    from ngame_account_mapping_agent     import NGameAccountMappingAgent
    from ngame_llm_analysis_agent        import NGameLLMAnalysisAgent
    from ngame_management_warning_agent  import NGameManagementWarningAgent
    from ngame_fraud_analysis_flow_manager import NGameFraudAnalysisFlowManager
    from datetime import datetime

    print("\n── Step 1: Churn comparison ──────────────────────────────────")
    churn_agent = NGameChurnComparisonAgent()
    comparison  = churn_agent.compare_churn_profiles(today_cpi_array, MATRIX_FILE)
    if not comparison['success']:
        raise RuntimeError(f"Churn comparison failed: {comparison['error']}")

    print("\n── Step 2: Rank differences ──────────────────────────────────")
    ranking = churn_agent.rank_differences(comparison['differences'])
    if not ranking['success']:
        raise RuntimeError(f"Ranking failed: {ranking['error']}")

    print("\n── Step 3: Top-3 anomaly identification ──────────────────────")
    anomaly_agent  = NGameAnomalyIdentificationAgent()
    anomaly_result = anomaly_agent.identify_top_anomalies(
        ranking['ranked_differences'], top_n=3
    )
    if not anomaly_result['success']:
        raise RuntimeError(f"Anomaly ID failed: {anomaly_result['error']}")

    for a in anomaly_result['top_anomalies']:
        print(f"  🚨 #{a['rank']}  φ{a['index']}  {a['transaction_type']}"
              f"  z={a.get('z_score', 0):.2f}  [{a['deviation_level']}]")

    print("\n── Step 4: Account mapping ───────────────────────────────────")
    account_agent  = NGameAccountMappingAgent()
    account_result = account_agent.map_to_chart_of_accounts(
        anomaly_result['top_anomalies']
    )
    if not account_result['success']:
        raise RuntimeError(f"Account mapping failed: {account_result['error']}")

    print("\n── Step 5: LLM misappropriation analysis ─────────────────────")
    llm_agent  = NGameLLMAnalysisAgent()
    llm_result = llm_agent.analyze_misappropriation_risks(
        account_result['account_mappings'],
        os.path.join(ROOT, "Asset_Misappropriation.ttl"),
    )
    if not llm_result['success']:
        raise RuntimeError(f"LLM analysis failed: {llm_result['error']}")

    print("\n── Step 6: Management warnings ───────────────────────────────")
    warning_agent  = NGameManagementWarningAgent()
    warning_result = warning_agent.generate_management_warnings(
        llm_result['matches'], llm_result['risk_assessments']
    )
    if not warning_result['success']:
        raise RuntimeError(f"Warning generation failed: {warning_result['error']}")

    print(f"\n  ✅ {len(warning_result['warnings'])} management warning(s) generated")

    # ── Assemble result identical to the real fraud analysis output ───────────
    result = {
        'success':          True,
        'execution_time':   datetime.now().isoformat(),
        'demo_scenario':    True,
        'injected_spikes':  {f"φ{i+1}_{n}": f"z≈{z}" for i, (n, z) in SPIKE_INDICES.items()},
        'today_cpi_array':  today_cpi_array,
        'comparison_result': comparison,
        'ranking_result':    ranking,
        'anomaly_result':    anomaly_result,
        'account_result':    account_result,
        'llm_result':        llm_result,
        'warning_result':    warning_result,
        'summary': {
            'total_transaction_types':        len(today_cpi_array),
            'top_anomalies_identified':       len(anomaly_result['top_anomalies']),
            'accounts_analyzed':              len(account_result['account_mappings']),
            'llm_matches_found':              len(llm_result['matches']),
            'management_warnings_generated':  len(warning_result['warnings']),
            'overall_risk_level':             warning_result.get('overall_risk_level', 'HIGH'),
        },
    }

    # ── Save artifacts so the dashboard picks them up ─────────────────────────
    mgr = NGameFraudAnalysisFlowManager()
    mgr.save_fraud_analysis_results(result)

    return result


def main():
    print("=" * 62)
    print("  🎯 NGAME DEMO SCENARIO — Controlled Anomaly Injection")
    print("=" * 62)

    if not os.path.exists(MATRIX_FILE):
        print(f"\n❌ Training matrix not found: {MATRIX_FILE}")
        print("   Run 'Run Training Day' first to create it.")
        return False

    print(f"\n📊 Loading baseline from {os.path.basename(MATRIX_FILE)} …")
    mu, sigma = load_baseline(MATRIX_FILE)
    print(f"   Loaded {len(mu)} transaction types.")

    print(f"\n💉 Injecting {len(SPIKE_INDICES)} controlled anomaly spike(s) [Vendors · ChartOfAccounts · Contractors] …")
    today_array = build_fraud_array(mu, sigma)

    print(f"\n🔍 Running fraud-analysis pipeline …")
    try:
        result = run_pipeline(today_array)
    except Exception as exc:
        print(f"\n❌ Pipeline error: {exc}")
        import traceback; traceback.print_exc()
        return False

    # ── Summary ───────────────────────────────────────────────────────────────
    s = result['summary']
    print("\n" + "=" * 62)
    print("  📋 DEMO SCENARIO COMPLETE")
    print("=" * 62)
    print(f"  Overall risk level   : {s['overall_risk_level']}")
    print(f"  Top anomalies found  : {s['top_anomalies_identified']}")
    print(f"  Management warnings  : {s['management_warnings_generated']}")
    print(f"  Injected spikes      : φ18 Vendors (z≈5.0) · φ16 ChartOfAccounts (z≈4.0) · φ14 Contractors (z≈3.5)")
    print("=" * 62)
    print("  ✅ Artifacts saved — refresh the Dashboard to see results.")
    print("=" * 62)
    return True


if __name__ == "__main__":
    os.chdir(ROOT)          # ensure relative paths resolve correctly
    success = main()
    sys.exit(0 if success else 1)
