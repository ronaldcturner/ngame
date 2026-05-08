#!/usr/bin/env python3
"""
Training Agent Excel Manager
Creates and manages Excel files with specific format:
- Column A: Transaction Type
- Column B: μ — Average CPI (mean of all training days)
- Column C: σ — Std Dev CPI (sample std dev of all training days)
- Column D+: Day 1, Day 2, Day 3, etc. (individual CPI values)
"""

import json
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

class TrainingAgentExcelManager:
    """Manages Excel files for Training Agent with specific column format."""
    
    def __init__(self, excel_file: str = "Training_Agent_Matrix.xlsx"):
        self.excel_file = excel_file
        self.workbook = None
        self.worksheet = None
        
    def create_initial_excel(self, baseline_data_file: str = "baseline_churn_matrix.json"):
        """Create initial Excel file from baseline data."""
        print("📊 Creating Initial Training Agent Excel")
        print("=" * 50)
        
        # Load baseline data
        with open(baseline_data_file, 'r') as f:
            baseline_data = json.load(f)
        
        # Create workbook
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Training Agent Matrix"
        
        # Get transaction types and CPI values
        transaction_types = baseline_data['transaction_types']
        cpi_values = baseline_data['baseline_cpi_values']
        
        # Create headers: A=Transaction Type, B=μ, C=σ, D=Day 1
        headers = ["Transaction Type", "μ (Mean CPI)", "σ (Std Dev CPI)", "Day 1"]
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for i, (transaction_type, cpi_value) in enumerate(zip(transaction_types, cpi_values), 2):
            # Transaction Type
            self.worksheet.cell(row=i, column=1, value=transaction_type)
            
            # μ — same as Day 1 on first day
            self.worksheet.cell(row=i, column=2, value=cpi_value)
            self.worksheet.cell(row=i, column=2).number_format = '0.000000'
            
            # σ — undefined with one data point; stored as 0.0
            self.worksheet.cell(row=i, column=3, value=0.0)
            self.worksheet.cell(row=i, column=3).number_format = '0.000000'
            
            # Day 1 value
            self.worksheet.cell(row=i, column=4, value=cpi_value)
            self.worksheet.cell(row=i, column=4).number_format = '0.000000'
        
        # Auto-fit columns
        self._auto_fit_columns()
        
        # Save workbook
        self.workbook.save(self.excel_file)
        
        print(f"✅ Initial Excel created: {self.excel_file}")
        print(f"📊 Transaction types: {len(transaction_types)}")
        print(f"📈 Columns: Transaction Type, μ (Mean CPI), σ (Std Dev CPI), Day 1")
        
        return self.excel_file
    
    def append_new_day(self, new_cpi_data_file: str = "Today_Churn_Matrix_Ready.json"):
        """Append a new day column with CPI values."""
        print("📈 Appending New Day Column")
        print("=" * 50)
        
        # Load existing workbook
        self.workbook = load_workbook(self.excel_file)
        self.worksheet = self.workbook.active
        
        # Load new CPI data
        with open(new_cpi_data_file, 'r') as f:
            new_cpi_data = json.load(f)
        
        # Extract similarity measures
        new_similarity_values = []
        new_similarity_labels = []
        
        for label, value in zip(new_cpi_data['pandas_series_ready']['index'], new_cpi_data['numpy_array']):
            if "_similarity" in label:
                transaction_type = label.replace("_similarity", "")
                new_similarity_values.append(value)
                new_similarity_labels.append(transaction_type)
        
        # Get current day number
        current_day = self._get_current_day_number()
        new_day = current_day + 1
        
        # Add new day header
        new_day_header = f"Day {new_day}"
        new_col = self.worksheet.max_column + 1
        self.worksheet.cell(row=1, column=new_col, value=new_day_header)
        self.worksheet.cell(row=1, column=new_col).font = Font(bold=True, color="FFFFFF")
        self.worksheet.cell(row=1, column=new_col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.worksheet.cell(row=1, column=new_col).alignment = Alignment(horizontal="center")
        
        # Add new day values
        for i, (transaction_type, cpi_value) in enumerate(zip(new_similarity_labels, new_similarity_values), 2):
            # Find the row for this transaction type
            for row in range(2, self.worksheet.max_row + 1):
                if self.worksheet.cell(row=row, column=1).value == transaction_type:
                    self.worksheet.cell(row=row, column=new_col, value=cpi_value)
                    self.worksheet.cell(row=row, column=new_col).number_format = '0.000000'
                    break
        
        # Update μ and σ columns
        self._update_average_cpi_coefficients()
        
        # Auto-fit columns
        self._auto_fit_columns()
        
        # Save updated workbook
        self.workbook.save(self.excel_file)
        
        print(f"✅ Added {new_day_header} column")
        print(f"📊 New day values added for {len(new_similarity_values)} transaction types")
        print(f"📈 μ and σ columns updated")
        
        return new_day
    
    def _get_current_day_number(self):
        """Get the current day number from the Excel file."""
        if not os.path.exists(self.excel_file):
            return 0
        
        # Load workbook to check current day
        temp_workbook = load_workbook(self.excel_file)
        temp_worksheet = temp_workbook.active
        
        # Count day columns (A=type, B=μ, C=σ; Day columns start at D=4)
        day_columns = 0
        for col in range(4, temp_worksheet.max_column + 1):
            header = temp_worksheet.cell(row=1, column=col).value
            if header and header.startswith("Day "):
                day_columns += 1
        
        return day_columns
    
    def _update_average_cpi_coefficients(self):
        """Update μ (col B) and σ (col C) for all rows from Day columns (col D onward)."""
        print("🔄 Updating μ and σ columns")
        
        # Day columns start at col 4 (A=type, B=μ, C=σ, D+=days)
        day_columns = []
        for col in range(4, self.worksheet.max_column + 1):
            header = self.worksheet.cell(row=1, column=col).value
            if header and header.startswith("Day "):
                day_columns.append(col)
        
        for row in range(2, self.worksheet.max_row + 1):
            cpi_values = []
            for col in day_columns:
                value = self.worksheet.cell(row=row, column=col).value
                if value is not None:
                    cpi_values.append(value)
            
            if cpi_values:
                mean_cpi = float(np.mean(cpi_values))
                # Sample std dev (ddof=1); falls back to 0.0 when only one day exists
                std_cpi = float(np.std(cpi_values, ddof=1)) if len(cpi_values) >= 2 else 0.0

                self.worksheet.cell(row=row, column=2, value=mean_cpi)
                self.worksheet.cell(row=row, column=2).number_format = '0.000000'

                self.worksheet.cell(row=row, column=3, value=std_cpi)
                self.worksheet.cell(row=row, column=3).number_format = '0.000000'
    
    def _auto_fit_columns(self):
        """Auto-fit column widths."""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def get_matrix_summary(self):
        """Get summary of the current matrix."""
        if not os.path.exists(self.excel_file):
            return None
        
        self.workbook = load_workbook(self.excel_file)
        self.worksheet = self.workbook.active
        
        # Get transaction types
        transaction_types = []
        for row in range(2, self.worksheet.max_row + 1):
            transaction_type = self.worksheet.cell(row=row, column=1).value
            if transaction_type:
                transaction_types.append(transaction_type)
        
        # Day columns start at col 4 (A=type, B=μ, C=σ, D+=days)
        day_columns = []
        for col in range(4, self.worksheet.max_column + 1):
            header = self.worksheet.cell(row=1, column=col).value
            if header and header.startswith("Day "):
                day_columns.append(col)
        
        # Read μ (col 2) and σ (col 3) for each transaction type row
        mean_cpi_values = []
        std_cpi_values = []
        for row in range(2, self.worksheet.max_row + 1):
            mean_val = self.worksheet.cell(row=row, column=2).value
            std_val  = self.worksheet.cell(row=row, column=3).value
            if mean_val is not None:
                mean_cpi_values.append(float(mean_val))
                std_cpi_values.append(float(std_val) if std_val is not None else 0.0)
        
        summary = {
            "excel_file": self.excel_file,
            "transaction_types": transaction_types,
            "total_transaction_types": len(transaction_types),
            "total_days": len(day_columns),
            "mean_cpi_values": mean_cpi_values,
            "std_cpi_values": std_cpi_values,
            # Cross-type summary statistics
            "mean_average_cpi": float(np.mean(mean_cpi_values)) if mean_cpi_values else 0.0,
            "std_average_cpi":  float(np.std(mean_cpi_values))  if mean_cpi_values else 0.0,
        }
        
        return summary

def create_training_agent_script():
    """Create a script that the Training Agent can run to append new days."""
    print("🤖 Creating Training Agent Script")
    print("=" * 50)
    
    training_script = '''#!/usr/bin/env python3
"""
Training Agent Script - Append New Day Column
This script is run by the Training Agent to append new CPI data.
"""

import sys
import os
from training_agent_excel_manager import TrainingAgentExcelManager

def main():
    """Main function for Training Agent to append new day."""
    print("🤖 Training Agent - Appending New Day")
    print("=" * 50)
    
    # Initialize Excel manager
    excel_manager = TrainingAgentExcelManager()
    
    # Check if Excel file exists
    if not os.path.exists(excel_manager.excel_file):
        print("❌ Excel file not found. Creating initial file...")
        excel_manager.create_initial_excel()
    
    # Append new day (assumes new CPI data is in Today_Churn_Matrix_Ready.json)
    new_day = excel_manager.append_new_day()
    
    # Get matrix summary
    summary = excel_manager.get_matrix_summary()
    
    if summary:
        print(f"\\n📊 MATRIX SUMMARY:")
        print(f"   Transaction types: {summary['total_transaction_types']}")
        print(f"   Total days: {summary['total_days']}")
        print(f"   Mean μ (across types): {summary['mean_average_cpi']:.6f}")
        print(f"   Std μ (across types):  {summary['std_average_cpi']:.6f}")
        
        print(f"\\n✅ Training Agent completed successfully!")
        print(f"📁 Excel file: {summary['excel_file']}")
        print(f"📈 New day added: Day {new_day}")
        
        return True
    else:
        print("❌ Failed to get matrix summary")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
'''
    
    with open("training_agent_append_day.py", 'w') as f:
        f.write(training_script)
    
    print("✅ Training agent script created: training_agent_append_day.py")
    return "training_agent_append_day.py"

def main():
    """Main function to demonstrate the Training Agent Excel system."""
    print("🚀 Training Agent Excel Manager")
    print("=" * 60)
    
    # Initialize Excel manager
    excel_manager = TrainingAgentExcelManager()
    
    # Create initial Excel file
    excel_file = excel_manager.create_initial_excel()
    
    # Create training agent script
    training_script = create_training_agent_script()
    
    # Get initial summary
    summary = excel_manager.get_matrix_summary()
    
    if summary:
        print(f"\n📊 INITIAL MATRIX SUMMARY:")
        print(f"   Excel file: {summary['excel_file']}")
        print(f"   Transaction types: {summary['total_transaction_types']}")
        print(f"   Total days: {summary['total_days']}")
        print(f"   Mean μ (across types): {summary['mean_average_cpi']:.6f}")
        print(f"   Std μ (across types):  {summary['std_average_cpi']:.6f}")
        
        print(f"\n🎯 TRAINING AGENT READY:")
        print(f"   1. Excel file created: {excel_file}")
        print(f"   2. Training script created: {training_script}")
        print(f"   3. Format: Transaction Type, μ (Mean CPI), σ (Std Dev CPI), Day 1")
        print(f"   4. Ready for new day appending")
        
        print(f"\n📈 NEXT STEPS:")
        print(f"   1. Run Phase 2 to generate new CPI data")
        print(f"   2. Run: python3 training_agent_append_day.py")
        print(f"   3. New day column will be added")
        print(f"   4. μ and σ columns will be updated")
        
        return True
    else:
        print("❌ Failed to create initial matrix")
        return False

if __name__ == "__main__":
    main()
