#!/usr/bin/env python3
"""Build Analysis Trends and Recent Activity payloads for the NGAME dashboard."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from ngame_transaction_types import CPI_ARRAY_SIZE

ACTIVITY_FILE = "ngame_dashboard_activity.json"
BUFFER_FILE = "NGAME_CPI_Rolling_Buffer.json"
MATRIX_FILE = "NGAME_Training_Matrix.xlsx"


def _try_load_json(path: Path) -> Optional[Any]:
    if not path.is_file():
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


def _churn_metrics_from_cpi_array(cpi_array: List[float]) -> Dict[str, float]:
    """
    Return peak and mean churn for one day's CPI column.

    peak — max(1 − CPI_k) across φ1–φ18 (worst category that day; best for chart visibility)
    mean — 1 − average(CPI) (smoother; often near zero when most categories are stable)
    """
    if not cpi_array:
        return {"peak_churn": 0.0, "mean_churn": 0.0, "mean_cpi": 1.0}
    floats = [float(v) for v in cpi_array]
    mean_cpi = sum(floats) / len(floats)
    per_type_churn = [max(0.0, 1.0 - v) for v in floats]
    return {
        "peak_churn": round(max(per_type_churn), 4),
        "mean_churn": round(max(0.0, 1.0 - mean_cpi), 4),
        "mean_cpi": round(mean_cpi, 4),
    }


def _churn_index_from_cpi_array(cpi_array: List[float]) -> float:
    """Peak category churn (used for chart and rolling buffer)."""
    return _churn_metrics_from_cpi_array(cpi_array)["peak_churn"]


def _training_matrix_trend_points(matrix_path: Path) -> List[Dict[str, Any]]:
    """One point per Day column in the training matrix."""
    if not matrix_path.is_file():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    points: List[Dict[str, Any]] = []
    try:
        wb = load_workbook(matrix_path, read_only=True, data_only=True)
        ws = wb.active
        max_row = min(ws.max_row or 0, 1 + CPI_ARRAY_SIZE)
        for col in range(4, (ws.max_column or 0) + 1):
            header = ws.cell(row=1, column=col).value
            label = str(header).strip() if header else f"Day {col - 3}"
            values: List[float] = []
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=col).value
                if cell is None:
                    continue
                try:
                    values.append(float(cell))
                except (TypeError, ValueError):
                    continue
            metrics = _churn_metrics_from_cpi_array(values) if values else {
                "peak_churn": 0.0,
                "mean_churn": 0.0,
                "mean_cpi": 1.0,
            }
            points.append(
                {
                    "label": label,
                    "churn_index": metrics["peak_churn"],
                    "mean_churn": metrics["mean_churn"],
                    "mean_cpi": metrics["mean_cpi"],
                    "empty_column": not values,
                    "source": "training_matrix",
                }
            )
        wb.close()
    except Exception:
        return []
    return points


def _rolling_buffer_trend_points(buffer_path: Path) -> List[Dict[str, Any]]:
    """Phase II rolling-buffer entries (fraud-check days)."""
    data = _try_load_json(buffer_path)
    if not isinstance(data, list):
        return []
    points: List[Dict[str, Any]] = []
    for i, entry in enumerate(data):
        if not isinstance(entry, dict):
            continue
        cpi = entry.get("cpi_array") or []
        if not cpi:
            continue
        date = entry.get("date") or entry.get("timestamp") or f"Run {i + 1}"
        label = str(date)[:10] if isinstance(date, str) else f"Run {i + 1}"
        points.append(
            {
                "label": label,
                "churn_index": _churn_index_from_cpi_array(cpi),
                "mean_cpi": round(sum(float(v) for v in cpi) / len(cpi), 4),
                "source": "rolling_buffer",
                "timestamp": entry.get("timestamp"),
            }
        )
    return points


def build_analysis_trends(
    root: Path,
    matrix_filename: str = MATRIX_FILE,
    buffer_filename: str = BUFFER_FILE,
) -> Dict[str, Any]:
    """
    Chart payload: daily peak category churn from training matrix + rolling buffer.
    """
    matrix_points = _training_matrix_trend_points(root / matrix_filename)
    buffer_points = _rolling_buffer_trend_points(root / buffer_filename)

    labels: List[str] = []
    churn_values: List[float] = []
    mean_churn_values: List[float] = []
    sources: List[str] = []
    empty_columns = 0

    for pt in matrix_points:
        labels.append(pt["label"])
        churn_values.append(pt["churn_index"])
        mean_churn_values.append(pt.get("mean_churn", pt["churn_index"]))
        sources.append(pt["source"])
        if pt.get("empty_column"):
            empty_columns += 1

    matrix_label_set = set(labels)
    for pt in buffer_points:
        lbl = pt["label"]
        if lbl in matrix_label_set and len(buffer_points) == 1:
            lbl = f"{lbl} (fraud check)"
        if lbl not in labels:
            labels.append(lbl)
            churn_values.append(pt["churn_index"])
            mean_churn_values.append(pt["churn_index"])
            sources.append(pt["source"])

    days_with_data = sum(1 for p in matrix_points if not p.get("empty_column"))

    return {
        "labels": labels,
        "churn_index": churn_values,
        "mean_churn_index": mean_churn_values,
        "sources": sources,
        "description": (
            "Peak category churn per day — highest (1 − CPI) across φ1–φ18. "
            "Spikes show which day had the wildest single category."
        ),
        "matrix_columns": len(matrix_points),
        "days_with_data": days_with_data,
        "empty_columns": empty_columns,
        "training_days": days_with_data,
        "buffer_days": len(buffer_points),
    }


def append_dashboard_activity(
    root: Path,
    event: Dict[str, Any],
    max_events: int = 25,
) -> None:
    """Persist a dashboard activity event (survives server restart)."""
    path = root / ACTIVITY_FILE
    events: List[Dict[str, Any]] = []
    existing = _try_load_json(path)
    if isinstance(existing, list):
        events = existing
    event.setdefault("timestamp", datetime.now().isoformat())
    events.append(event)
    events = events[-max_events:]
    try:
        with path.open("w", encoding="utf-8") as f:
            json.dump(events, f, indent=2)
    except OSError:
        pass


def _load_persisted_activity(root: Path) -> List[Dict[str, Any]]:
    data = _try_load_json(root / ACTIVITY_FILE)
    return data if isinstance(data, list) else []


def build_recent_activity(
    root: Path,
    fraud: Optional[Dict[str, Any]] = None,
    training_status: Optional[Dict[str, Any]] = None,
    fraud_summary: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """Recent pipeline events for the activity feed, newest first."""
    events: List[Dict[str, Any]] = []

    for item in _load_persisted_activity(root):
        if isinstance(item, dict) and item.get("type"):
            events.append(
                {
                    "type": item.get("type"),
                    "message": item.get("message") or "",
                    "timestamp": item.get("timestamp"),
                    "status": item.get("status") or "info",
                }
            )

    fraud = fraud or {}
    summary = fraud_summary or fraud.get("summary") or {}
    execution_time = fraud.get("execution_time")
    if execution_time:
        top_n = summary.get("top_anomalies_identified", 0)
        risk = summary.get("overall_risk_level", "UNKNOWN")
        warnings = summary.get("management_warnings_generated", 0)
        dollar_hi = summary.get("dollar_alarm_high_count")
        if dollar_hi is None:
            dollar_hi = 0
        events.append(
            {
                "type": "Fraud check",
                "message": (
                    f"Risk {risk}; {top_n} top anomalies; "
                    f"{warnings} management warning(s)"
                ),
                "timestamp": execution_time,
                "status": "success" if risk == "LOW" else "warning",
            }
        )

    buffer = _try_load_json(root / BUFFER_FILE)
    if isinstance(buffer, list):
        for entry in buffer[-5:]:
            if not isinstance(entry, dict):
                continue
            ts = entry.get("timestamp") or entry.get("date")
            if not ts:
                continue
            events.append(
                {
                    "type": "Rolling buffer",
                    "message": "CPI snapshot stored for multi-day pattern",
                    "timestamp": ts,
                    "status": "info",
                }
            )

    if training_status:
        days = training_status.get("days_recorded") or 0
        target = training_status.get("target_days") or 30
        if training_status.get("matrix_exists") and days > 0:
            msg = (
                f"Training complete ({days}/{target})"
                if training_status.get("training_complete")
                else f"Matrix: {days} of {target} business days recorded"
            )
            matrix_path = root / (training_status.get("matrix_file") or MATRIX_FILE)
            mtime = (
                datetime.fromtimestamp(matrix_path.stat().st_mtime).isoformat()
                if matrix_path.is_file()
                else None
            )
            events.append(
                {
                    "type": "Training matrix",
                    "message": msg,
                    "timestamp": mtime,
                    "status": "success" if training_status.get("training_complete") else "info",
                }
            )

    # De-duplicate by type+timestamp, sort newest first
    seen = set()
    unique: List[Dict[str, Any]] = []
    for ev in sorted(
        events,
        key=lambda e: e.get("timestamp") or "",
        reverse=True,
    ):
        key = (ev.get("type"), ev.get("timestamp"), ev.get("message"))
        if key in seen:
            continue
        seen.add(key)
        unique.append(ev)

    return unique[:10]
