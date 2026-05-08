#!/usr/bin/env python3
"""
NGAME Training Flow Launcher
Ensures correct Python environment and runs the training flow.

Usage
-----
Demo mode — first call of the live demonstration (Segment 3):
    python3 run_training_flow.py --demo

    Backs up any complete 30-day matrix to NGAME_Training_Matrix_SAVED.xlsx,
    then starts an interactive loop.  Answer 'y' to add each day in sequence;
    answer 'n' (or just press Enter) to pause so students can inspect the
    matrix, then run again (no flag needed) to resume.

Resume / continue adding days (Day 2, 3, … after --demo):
    python3 run_training_flow.py

    Detected saved matrix triggers automatic replay — no flag needed.
    Answers 'y' to keep adding columns, 'n' to pause.

Reset — wipe the demo matrix and start the demo from Day 1 again:
    python3 run_training_flow.py --reset

    Deletes the current demo matrix only; the saved 30-day matrix is kept
    intact.  Run --demo again afterwards to replay from Day 1.

Restore — after the demo, before Time Warp (Segment 4):
    python3 run_training_flow.py --restore

    Restores the saved 30-day matrix and removes the demo matrix so the
    pipeline is ready for fraud analysis.
"""

import sys
import os
import subprocess

MATRIX_FILE = "NGAME_Training_Matrix.xlsx"
SAVED_MATRIX_FILE = "NGAME_Training_Matrix_SAVED.xlsx"


# ── Environment helpers ───────────────────────────────────────────────────────

def check_environment():
    print("🔍 Checking Python environment...")
    print(f"Python executable: {sys.executable}")
    print(f"Python version: {sys.version}")
    try:
        import openpyxl
        print(f"✅ openpyxl version: {openpyxl.__version__}")
        return True
    except ImportError as e:
        print(f"❌ openpyxl not available: {e}")
        return False


def install_dependencies():
    print("📦 Installing dependencies...")
    try:
        result = subprocess.run(
            [sys.executable, '-m', 'pip', 'install', 'openpyxl'],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode == 0:
            print("✅ openpyxl installed successfully")
            return True
        print(f"❌ Failed to install openpyxl: {result.stderr}")
        return False
    except Exception as e:
        print(f"❌ Error installing dependencies: {e}")
        return False


# ── Matrix helpers ────────────────────────────────────────────────────────────

def _count_training_days(matrix_path: str) -> int:
    """Return the number of Day columns in the matrix, or 0 on any error."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(matrix_path, read_only=True)
        ws = wb.active
        days = max(ws.max_column - 3, 0)   # A=type, B=μ, C=σ → Day cols start at 4
        wb.close()
        return days
    except Exception:
        return 0


def _read_day_column(matrix_path: str, day_number: int) -> list:
    """
    Read the CPI values for a specific training day from a matrix file.
    Day columns start at column 4 (A=1, B=2, C=3, Day1=4, Day2=5, …).
    """
    from openpyxl import load_workbook
    col = day_number + 3
    wb  = load_workbook(matrix_path, read_only=True)
    ws  = wb.active
    cpi = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col).value
        cpi.append(float(val) if val is not None else 0.0)
    wb.close()
    return cpi


# ── Demo backup / restore ─────────────────────────────────────────────────────

def demo_backup():
    """
    Back up any existing complete matrix so a fresh Day 1 can be shown live.
    Called when --demo flag is used.
    """
    if not os.path.exists(MATRIX_FILE):
        print("ℹ️  No existing matrix found — proceeding straight to Day 1.")
        return True

    days = _count_training_days(MATRIX_FILE)

    if days >= 30:
        if os.path.exists(SAVED_MATRIX_FILE):
            os.remove(SAVED_MATRIX_FILE)
        os.rename(MATRIX_FILE, SAVED_MATRIX_FILE)
        print(f"✅ 30-day matrix backed up → {SAVED_MATRIX_FILE}")
        print(f"   Fresh slate — Day 1 values will be replayed from saved matrix.")
    elif days > 0:
        print(f"ℹ️  Existing matrix has only {days} day(s) — already a demo matrix.")
        print(f"   No backup needed; proceeding to add the next day.")
    else:
        print("ℹ️  Matrix file exists but has no Day columns — proceeding to Day 1.")

    return True


def demo_restore():
    """
    Restore the 30-day backup and discard the demo matrix.
    Called when --restore flag is used.
    """
    if not os.path.exists(SAVED_MATRIX_FILE):
        print(f"⚠️  No saved matrix found at {SAVED_MATRIX_FILE}")
        print(f"   Nothing to restore.")
        return False

    if os.path.exists(MATRIX_FILE):
        demo_days = _count_training_days(MATRIX_FILE)
        os.remove(MATRIX_FILE)
        print(f"🗑️  Removed {demo_days}-day demo matrix.")

    os.rename(SAVED_MATRIX_FILE, MATRIX_FILE)
    restored_days = _count_training_days(MATRIX_FILE)
    print(f"✅ {restored_days}-day training matrix restored → {MATRIX_FILE}")
    print(f"   Pipeline is ready for fraud analysis.")
    return True


def demo_reset():
    """
    Delete only the current demo matrix so the demo can be replayed from
    Day 1.  The saved 30-day matrix is left untouched.
    Called when --reset flag is used.
    """
    if not os.path.exists(MATRIX_FILE):
        print(f"ℹ️  No demo matrix found at {MATRIX_FILE} — nothing to reset.")
    else:
        days = _count_training_days(MATRIX_FILE)
        os.remove(MATRIX_FILE)
        print(f"🗑️  {days}-day demo matrix removed.")

    if os.path.exists(SAVED_MATRIX_FILE):
        saved_days = _count_training_days(SAVED_MATRIX_FILE)
        print(f"   Saved {saved_days}-day matrix is intact.")
        print(f"   Run --demo again to replay from Day 1.")
    else:
        print(f"   No saved matrix found — run --demo to begin.")
    return True


# ── Demo day replay ───────────────────────────────────────────────────────────

def run_demo_day_replay(next_day: int) -> dict:
    """
    Write pre-stored CPI values from the saved 30-day matrix into the live
    demo matrix — no QuickBooks connection or CPI calculation needed.

    This makes Day 1 and Day 2 of the live demo look identical to what a
    real training run would have produced.
    """
    print(f"📋 Demo replay: reading Day {next_day} values from saved 30-day matrix …")

    try:
        cpi_array = _read_day_column(SAVED_MATRIX_FILE, next_day)
        if not cpi_array:
            return {'success': False, 'error': 'No CPI values read from saved matrix'}

        from ngame_matrix_management_agent import NGameMatrixManagementAgent
        agent  = NGameMatrixManagementAgent()
        result = agent.update_training_matrix(next_day, cpi_array)

        if result.get('success'):
            print(f"✅ Day {next_day} values replayed from saved matrix "
                  f"({len(cpi_array)} transaction types).")
        return result

    except Exception as e:
        return {'success': False, 'error': str(e)}


# ── Training flow (normal and demo) ──────────────────────────────────────────

def run_training_flow(demo_mode: bool = False) -> bool:
    """
    Interactive loop: adds training days one at a time.

    After each day the user is asked whether to continue immediately (y) or
    pause (n / Enter) so students can open and inspect the matrix.  Running
    the script again resumes from where it left off.

    In demo_mode (or when the saved matrix exists), pre-stored values from
    NGAME_Training_Matrix_SAVED.xlsx are used instead of a live extraction.
    """
    print("🚀 Starting NGAME Training Flow...")

    try:
        from ngame_training_flow_manager import NGameTrainingFlowManager
        training_manager = NGameTrainingFlowManager()

        while True:
            status        = training_manager.get_training_status()
            days_recorded = status.get('training_days', 0)
            target_days   = status['target_days']
            next_day      = days_recorded + 1

            print(f"\n📊 Training Status:")
            print(f"   Matrix exists    : {status['matrix_exists']}")
            print(f"   Days recorded    : {days_recorded} of {target_days}")
            if status['matrix_exists']:
                print(f"   Matrix dimensions: {status.get('matrix_dimensions', 'N/A')}")
                print(f"   Training complete: {status.get('training_complete', False)}")

            if status.get('training_complete', False):
                print(f"\n✅ Training period complete — ready for fraud analysis.")
                print(f"   (To re-run the live demo, use: python3 run_training_flow.py --demo)")
                return True

            print(f"\n❓ Add Day {next_day} of {target_days} to the training matrix? (y/n): ", end="", flush=True)
            response = input().lower().strip()
            if response not in ['y', 'yes']:
                print(f"\n⏹️  Paused at Day {days_recorded}.")
                print(f"   Open the matrix to show students, then run this script again to continue.")
                return True

            use_replay = demo_mode or os.path.exists(SAVED_MATRIX_FILE)
            if use_replay and not os.path.exists(SAVED_MATRIX_FILE):
                print(f"⚠️  Demo replay requested but {SAVED_MATRIX_FILE} not found.")
                print(f"   Falling back to live extraction.")
                use_replay = False

            if use_replay:
                print(f"\n📅 Adding Day {next_day} of {target_days} (demo replay) …")
                result = run_demo_day_replay(next_day)
            else:
                print(f"\n📅 Adding Day {next_day} of {target_days} …")
                result = training_manager.execute_training_period(next_day, next_day)

            days_remaining = target_days - next_day
            if result.get('success'):
                if days_remaining > 0:
                    print(f"\n✅ Day {next_day} recorded — {days_remaining} day(s) remaining.")
                    if not use_replay:
                        print(f"   (Live mode — return tomorrow to add Day {next_day + 1}.)")
                        return True
                    # Loop back: ask about next day immediately
                else:
                    print(f"\n🎉 All {target_days} training days collected!")
                    print(f"   Matrix is ready for fraud analysis.")
                    return True
            else:
                print(f"\n❌ Day {next_day} failed: {result.get('error', 'see errors above')}")
                return False

    except ImportError as e:
        print(f"❌ Import error: {e}")
        return False
    except Exception as e:
        print(f"❌ Error running training flow: {e}")
        return False


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    args       = sys.argv[1:]
    is_demo    = '--demo'    in args
    is_restore = '--restore' in args
    is_reset   = '--reset'   in args

    print("🚀 NGAME Training Flow Launcher")
    print("=" * 50)

    # ── Restore mode ──────────────────────────────────────────────────────────
    if is_restore:
        print("🔄 Restore mode: returning 30-day matrix for Time Warp segment\n")
        success = demo_restore()
        print(f"\n{'✅ Restore complete!' if success else '❌ Restore failed'}")
        return success

    # ── Reset mode ────────────────────────────────────────────────────────────
    if is_reset:
        print("🔄 Reset mode: clearing demo matrix to replay from Day 1\n")
        success = demo_reset()
        print(f"\n{'✅ Reset complete!' if success else '❌ Reset failed'}")
        return success

    # ── Environment check ─────────────────────────────────────────────────────
    if not check_environment():
        print("\n🔧 Attempting to fix environment...")
        if not install_dependencies():
            print("\n❌ Could not fix environment. Please check your Python setup.")
            return False

    # ── Demo mode: back up matrix, then start interactive loop from Day 1 ─────
    if is_demo:
        print("🎭 Demo mode: backing up pre-built matrix and replaying stored values\n")
        demo_backup()
        print()

    # ── Training flow (live or replay) ────────────────────────────────────────
    success = run_training_flow(demo_mode=is_demo)
    print(f"\n{'✅ Launcher completed successfully!' if success else '❌ Launcher failed'}")
    return success


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
