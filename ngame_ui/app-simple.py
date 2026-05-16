#!/usr/bin/env python3
"""
NGAME Web Application (Simplified Version)
A simplified version that works without WebSocket dependencies for easier installation.
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
import json
import os
import logging
from datetime import datetime, timedelta
import threading
import time
import subprocess
import sys as _sys
from pathlib import Path
from typing import Any, Dict, Optional, Tuple, List

# Import NGAME components (optional - will work without them)
try:
    import sys
    sys.path.append('..')
    from quickbooks_chart_of_accounts_extractor import ChartOfAccountsExtractor, OntologyConverter
    from quickbooks_chart_of_accounts_creator import OntologyChartExtractor, ChartOfAccountsCreator
    from ngame_ontology_workflow import NGAMEOntologyWorkflow
    NGAME_AVAILABLE = True
except ImportError:
    NGAME_AVAILABLE = False
    print("⚠️  NGAME components not available. Running in demo mode.")

app = Flask(__name__)
app.config['SECRET_KEY'] = 'ngame-secret-key-2024'

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent

TRAINING_MATRIX_FILE = "NGAME_Training_Matrix.xlsx"
TRAINING_TARGET_DAYS = 30
FRP_MODE = os.environ.get("NGAME_UI_FRP_MODE", "1").lower() not in ("0", "false", "no")


def _count_training_days(matrix_path: Path) -> int:
    """Day columns start at column 4 (A=type, B=μ, C=σ)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(matrix_path, read_only=True)
        ws = wb.active
        days = max(ws.max_column - 3, 0)
        wb.close()
        return days
    except Exception:
        return 0


def _get_training_status() -> Dict[str, Any]:
    matrix_path = _repo_root() / TRAINING_MATRIX_FILE
    if not matrix_path.exists():
        return {
            "matrix_exists": False,
            "days_recorded": 0,
            "target_days": TRAINING_TARGET_DAYS,
            "days_remaining": TRAINING_TARGET_DAYS,
            "training_complete": False,
            "phase": "training",
        }
    days = _count_training_days(matrix_path)
    complete = days >= TRAINING_TARGET_DAYS
    return {
        "matrix_exists": True,
        "days_recorded": days,
        "target_days": TRAINING_TARGET_DAYS,
        "days_remaining": max(0, TRAINING_TARGET_DAYS - days),
        "training_complete": complete,
        "phase": "fraud_analysis" if complete else "training",
    }

def _try_load_json(path: Path) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    try:
        if not path.exists():
            return None, None
        with path.open('r', encoding='utf-8') as f:
            return json.load(f), None
    except Exception as e:
        return None, f"{path}: {e}"

def _load_latest_artifacts() -> Dict[str, Any]:
    """
    Consolidate key NGAME run artifacts for the UI.

    Priority is root-level artifacts, then GENERATED_FILES fallbacks.
    """
    root = _repo_root()
    fallbacks = root / "GENERATED_FILES"

    candidates: Dict[str, List[Path]] = {
        "management_dashboard": [
            root / "management_dashboard.json",
            fallbacks / "management_dashboard.json",
        ],
        "fraud_analysis": [
            root / "NGAME_Fraud_Analysis_readable_truly_clean.json",
            root / "NGAME_Fraud_Analysis_readable_clean.json",
            root / "NGAME_Fraud_Analysis_readable.json",
            root / "NGAME_Fraud_Analysis.json",
            fallbacks / "NGAME_Fraud_Analysis_readable_truly_clean.json",
            fallbacks / "NGAME_Fraud_Analysis_readable_clean.json",
            fallbacks / "NGAME_Fraud_Analysis_readable.json",
            fallbacks / "NGAME_Fraud_Analysis.json",
        ],
    }

    errors: List[str] = []
    loaded: Dict[str, Any] = {}
    sources: Dict[str, str] = {}

    for key, paths in candidates.items():
        for p in paths:
            data, err = _try_load_json(p)
            if err:
                errors.append(err)
            if data is not None:
                loaded[key] = data
                sources[key] = str(p)
                break

    management_dashboard = loaded.get("management_dashboard") or {}
    fraud = loaded.get("fraud_analysis") or {}

    # Extract consolidated view
    management_summary = (management_dashboard.get("management_summary") or
                          fraud.get("warning_result", {}).get("management_summary") or {})
    warnings = management_dashboard.get("warnings") or fraud.get("warning_result", {}).get("warnings") or []
    overall_risk = (management_summary.get("overall_risk_level") or
                    fraud.get("summary", {}).get("overall_risk_level") or
                    fraud.get("warning_result", {}).get("overall_risk_level") or
                    "UNKNOWN")

    top_anomalies = (fraud.get("anomaly_result", {}) or {}).get("top_anomalies") or []
    fraud_summary = fraud.get("summary") or {}
    sequence_result = fraud.get("sequence_result")
    cc_watch_result = fraud.get("cc_watch_result") or {}

    sequence_active = bool(fraud_summary.get("sequence_active"))
    sequence_window = fraud_summary.get("sequence_window")
    sequence_high = int(fraud_summary.get("sequence_high_count") or 0)
    sequence_medium = int(fraud_summary.get("sequence_medium_count") or 0)
    sequence_top: List[Dict[str, Any]] = []
    if sequence_result and sequence_result.get("comparison", {}).get("success"):
        comp = sequence_result["comparison"]
        ranked = comp.get("ranked_differences") or comp.get("differences") or []
        sequence_top = sorted(
            ranked,
            key=lambda x: float(x.get("abs_z_score") or x.get("z_score") or 0),
            reverse=True,
        )[:5]

    cc_summary = cc_watch_result.get("summary") or {}
    cc_flagged = cc_watch_result.get("flagged_transactions") or []

    # Last-updated timestamps (best effort)
    last_updated = management_dashboard.get("last_updated") or management_dashboard.get("management_summary", {}).get("summary_timestamp")
    execution_time = fraud.get("execution_time")

    return {
        "success": True,
        "sources": sources,
        "errors": errors[-10:],  # keep response small
        "training_status": _get_training_status(),
        "summary": {
            "overall_risk_level": overall_risk,
            "warnings_count": len(warnings),
            "top_anomalies_count": len(top_anomalies),
            "last_updated": last_updated,
            "execution_time": execution_time,
            "sequence_active": sequence_active,
            "sequence_window": sequence_window,
            "sequence_high_count": sequence_high,
            "sequence_medium_count": sequence_medium,
            "cc_watch_flagged": int(cc_summary.get("total_flagged") or 0),
            "cc_watch_highest_risk": cc_summary.get("highest_risk_level", "CLEAR"),
        },
        "management_dashboard": {
            "management_summary": management_summary,
            "warnings": warnings[:50],
            "dashboard_status": management_dashboard.get("dashboard_status", "UNKNOWN"),
            "last_updated": last_updated,
        },
        "fraud_analysis": {
            "execution_time": execution_time,
            "anomaly_result": {
                "top_anomalies": top_anomalies[:10],
            },
            "llm_result": fraud.get("llm_result", {}),
            "summary": fraud_summary,
            "sequence_result": sequence_result,
            "sequence_top": sequence_top,
            "cc_watch_result": cc_watch_result,
            "cc_flagged": cc_flagged[:10],
        },
    }

class NGAMEState:
    """Global state management for NGAME application."""
    
    def __init__(self):
        self.workflow = None
        self.current_analysis = None
        self.analysis_history = []
        self.quickbooks_connected = False
        self.ontology_loaded = False
        self.active_alerts = []
        # ── Operation runner state ──────────────────────────────────────
        self.operation_running = False
        self.operation_name    = ""
        self.operation_log: List[str] = []
        self._operation_lock   = threading.Lock()
        
    def initialize_workflow(self):
        """Initialize the NGAME ontology workflow."""
        if not NGAME_AVAILABLE:
            return False
            
        try:
            self.workflow = NGAMEOntologyWorkflow()
            return True
        except Exception as e:
            logger.error(f"Error initializing workflow: {e}")
            return False
    
    def add_analysis(self, analysis_data):
        """Add analysis to history."""
        analysis_data['timestamp'] = datetime.now().isoformat()
        analysis_data['id'] = len(self.analysis_history) + 1
        self.analysis_history.append(analysis_data)
        
        # Keep only last 50 analyses
        if len(self.analysis_history) > 50:
            self.analysis_history = self.analysis_history[-50:]
    
    def get_dashboard_stats(self):
        """Get dashboard statistics."""
        return {
            'total_analyses': len(self.analysis_history),
            'active_alerts': len(self.active_alerts),
            'quickbooks_connected': self.quickbooks_connected,
            'ontology_loaded': self.ontology_loaded,
            'recent_analyses': self.analysis_history[-5:] if self.analysis_history else [],
            'ngame_available': NGAME_AVAILABLE
        }

# Initialize global state
ngame_state = NGAMEState()


@app.context_processor
def inject_ui_config():
    return {
        "frp_mode": FRP_MODE,
        "training_status": _get_training_status(),
    }


@app.route('/')
def index():
    """Main dashboard page."""
    return render_template('dashboard.html', stats=ngame_state.get_dashboard_stats())

@app.route('/dashboard')
def dashboard():
    """Dashboard with real-time updates."""
    return render_template('dashboard.html', stats=ngame_state.get_dashboard_stats())

@app.route('/dashboard-debug')
def dashboard_debug():
    """Debug dashboard showing raw consolidated payload."""
    return render_template('dashboard-debug.html')

@app.route('/analysis')
def analysis():
    """Analysis page for running anomaly detection."""
    return render_template('analysis.html')

@app.route('/ontology')
def ontology():
    """Ontology management page."""
    return render_template('ontology.html')

@app.route('/quickbooks')
def quickbooks():
    """QuickBooks integration page."""
    return render_template('quickbooks.html')

@app.route('/alerts')
def alerts():
    """Alerts and notifications page."""
    return render_template('alerts.html', alerts=ngame_state.active_alerts)

@app.route('/reports')
def reports():
    """Reports and analytics page."""
    return render_template('reports.html', history=ngame_state.analysis_history)

@app.route('/settings')
def settings():
    """Settings and configuration page."""
    return render_template('settings.html')

# API Routes

@app.route('/api/initialize', methods=['POST'])
def api_initialize():
    """Initialize NGAME system."""
    try:
        if not NGAME_AVAILABLE:
            return jsonify({
                'success': False, 
                'message': 'NGAME components not available. Running in demo mode.'
            })
            
        success = ngame_state.initialize_workflow()
        if success:
            return jsonify({'success': True, 'message': 'NGAME system initialized successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to initialize NGAME system'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/load-ontology', methods=['POST'])
def api_load_ontology():
    """Load ontology file."""
    try:
        if not NGAME_AVAILABLE:
            return jsonify({
                'success': False, 
                'message': 'NGAME components not available. Running in demo mode.'
            })
            
        data = request.get_json()
        ontology_path = data.get('ontology_path')
        
        if not ontology_path or not os.path.exists(ontology_path):
            return jsonify({'success': False, 'message': 'Ontology file not found'})
        
        # Load ontology using NGAME workflow
        if ngame_state.workflow:
            ngame_state.workflow.load_ontology(ontology_path)
            ngame_state.ontology_loaded = True
            return jsonify({'success': True, 'message': 'Ontology loaded successfully'})
        else:
            return jsonify({'success': False, 'message': 'NGAME workflow not initialized'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/run-analysis', methods=['POST'])
def api_run_analysis():
    """Run anomaly detection analysis."""
    try:
        data = request.get_json()
        analysis_type = data.get('type', 'comprehensive')
        
        # Simulate analysis for demo mode
        if not NGAME_AVAILABLE:
            # Create demo results
            demo_results = {
                'analysis_type': analysis_type,
                'timestamp': datetime.now().isoformat(),
                'duration': '2.3 seconds',
                'anomalies': [
                    {
                        'id': 1,
                        'title': 'High Balance Account Detected',
                        'description': 'Account "Checking Account" has unusually high balance',
                        'severity': 'medium',
                        'type': 'statistical_anomaly',
                        'timestamp': datetime.now().isoformat()
                    },
                    {
                        'id': 2,
                        'title': 'Duplicate Account Names',
                        'description': 'Multiple accounts with similar names detected',
                        'severity': 'low',
                        'type': 'data_quality',
                        'timestamp': datetime.now().isoformat()
                    }
                ],
                'summary': {
                    'total_anomalies': 2,
                    'high_severity': 0,
                    'medium_severity': 1,
                    'low_severity': 1
                }
            }
            
            ngame_state.current_analysis = demo_results
            ngame_state.add_analysis(demo_results)
            
            return jsonify({
                'success': True, 
                'message': 'Demo analysis completed',
                'results': demo_results
            })
        
        # Start analysis in background
        def run_analysis():
            try:
                if ngame_state.workflow:
                    results = ngame_state.workflow.run_ontology_analysis(data)
                    ngame_state.current_analysis = results
                    ngame_state.add_analysis(results)
                else:
                    logger.error("NGAME workflow not initialized")
            except Exception as e:
                logger.error(f"Error running analysis: {e}")
        
        # Run analysis in background thread
        thread = threading.Thread(target=run_analysis)
        thread.start()
        
        return jsonify({'success': True, 'message': 'Analysis started'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/extract-chart-of-accounts', methods=['POST'])
def api_extract_chart_of_accounts():
    """Extract chart of accounts from QuickBooks."""
    try:
        if not NGAME_AVAILABLE:
            # Return demo data
            return jsonify({
                'success': True,
                'message': 'Demo: Chart of accounts extracted successfully',
                'output_path': 'demo_chart_of_accounts.ttl',
                'account_count': 15
            })
            
        data = request.get_json()
        
        # Extract chart of accounts
        extractor = ChartOfAccountsExtractor(None)  # Will need API client
        chart_of_accounts = extractor.extract_chart_of_accounts()
        
        # Convert to ontology
        converter = OntologyConverter()
        ontology_graph = converter.convert_to_ontology(chart_of_accounts)
        
        # Save ontology
        output_path = f"output/chart_of_accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.ttl"
        converter.save_ontology(output_path)
        
        return jsonify({
            'success': True,
            'message': 'Chart of accounts extracted successfully',
            'output_path': output_path,
            'account_count': len(chart_of_accounts.accounts)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/create-chart-of-accounts', methods=['POST'])
def api_create_chart_of_accounts():
    """Create chart of accounts in QuickBooks from ontology."""
    try:
        if not NGAME_AVAILABLE:
            # Return demo data
            return jsonify({
                'success': True,
                'message': 'Demo: Chart of accounts creation preview',
                'result': {
                    'accounts_to_create': 8,
                    'accounts_to_skip': 2,
                    'dry_run': True
                }
            })
            
        data = request.get_json()
        ontology_path = data.get('ontology_path')
        
        # Extract from ontology
        extractor = OntologyChartExtractor()
        extractor.load_ontology(ontology_path)
        template = extractor.extract_chart_of_accounts()
        
        # Create in QuickBooks (dry run for safety)
        creator = ChartOfAccountsCreator(None)  # Will need API client
        result = creator.create_chart_of_accounts(template, dry_run=True)
        
        return jsonify({
            'success': True,
            'message': 'Chart of accounts creation preview',
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/dashboard-stats')
def api_dashboard_stats():
    """Get dashboard statistics."""
    return jsonify(ngame_state.get_dashboard_stats())

@app.route('/api/consolidated-dashboard')
def api_consolidated_dashboard():
    """Get consolidated dashboard data from NGAME run artifacts."""
    return jsonify(_load_latest_artifacts())

@app.route('/api/socketio-enabled')
def api_socketio_enabled():
    """Whether this server supports Socket.IO endpoints."""
    return jsonify({"enabled": False})

@app.route('/api/analysis-history')
def api_analysis_history():
    """Get analysis history."""
    return jsonify(ngame_state.analysis_history)

@app.route('/api/active-alerts')
def api_active_alerts():
    """Get active alerts."""
    return jsonify(ngame_state.active_alerts)

# ── Operation runner helpers ──────────────────────────────────────────────────

def _stream_script(script_name: str, stdin_input: str = "y\n"):
    """
    Run a top-level NGAME script as a subprocess, stream its output into
    ngame_state.operation_log, then mark operation_running = False.
    """
    root = _repo_root()
    script = root / script_name

    ngame_state.operation_log = [f"▶ Starting {script_name} …", ""]

    try:
        proc = subprocess.Popen(
            [_sys.executable, "-u", str(script)],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            stdin=subprocess.PIPE,
            cwd=str(root),
            text=True,
        )
        # Answer any y/n prompts automatically
        try:
            proc.stdin.write(stdin_input)
            proc.stdin.close()
        except Exception:
            pass

        for raw_line in proc.stdout:
            line = raw_line.rstrip()
            with ngame_state._operation_lock:
                ngame_state.operation_log.append(line)
                if len(ngame_state.operation_log) > 300:
                    ngame_state.operation_log = ngame_state.operation_log[-300:]

        proc.wait()
        exit_msg = "✅ Completed (exit 0)" if proc.returncode == 0 else f"⚠️  Exited with code {proc.returncode}"
        with ngame_state._operation_lock:
            ngame_state.operation_log.append("")
            ngame_state.operation_log.append(exit_msg)

    except Exception as exc:
        with ngame_state._operation_lock:
            ngame_state.operation_log.append(f"❌ Error: {exc}")
    finally:
        ngame_state.operation_running = False


@app.route('/api/training-status')
def api_training_status():
    """Training matrix progress for dashboard phase guidance."""
    return jsonify(_get_training_status())


@app.route('/api/ui-config')
def api_ui_config():
    return jsonify({"frp_mode": FRP_MODE, "training_status": _get_training_status()})


@app.route('/api/run-daily', methods=['POST'])
def api_run_daily():
    """Run training or fraud analysis based on current training matrix state."""
    status = _get_training_status()
    if status["phase"] == "training":
        return api_run_training()
    return api_run_fraud_analysis()


@app.route('/api/run-training', methods=['POST'])
def api_run_training():
    """Launch one training day (run_training_flow.py) in the background."""
    with ngame_state._operation_lock:
        if ngame_state.operation_running:
            return jsonify({'success': False, 'message': 'An operation is already running.'})
        ngame_state.operation_running = True
        ngame_state.operation_name    = "Training Day"
        ngame_state.operation_log     = []

    t = threading.Thread(
        target=_stream_script,
        args=("run_training_flow.py", "y\n"),
        daemon=True,
    )
    t.start()
    return jsonify({'success': True, 'message': 'Training run started.'})


@app.route('/api/run-fraud-analysis', methods=['POST'])
def api_run_fraud_analysis():
    """Launch daily fraud / churn analysis (run_fraud_analysis.py) in the background."""
    with ngame_state._operation_lock:
        if ngame_state.operation_running:
            return jsonify({'success': False, 'message': 'An operation is already running.'})
        ngame_state.operation_running = True
        ngame_state.operation_name    = "Daily Churn Analysis"
        ngame_state.operation_log     = []

    t = threading.Thread(
        target=_stream_script,
        args=("run_fraud_analysis.py", "y\n"),
        daemon=True,
    )
    t.start()
    return jsonify({'success': True, 'message': 'Fraud analysis started.'})


@app.route('/api/run-demo-scenario', methods=['POST'])
def api_run_demo_scenario():
    """Launch the controlled demo anomaly scenario (run_demo_scenario.py)."""
    with ngame_state._operation_lock:
        if ngame_state.operation_running:
            return jsonify({'success': False, 'message': 'An operation is already running.'})
        ngame_state.operation_running = True
        ngame_state.operation_name    = "Demo Scenario"
        ngame_state.operation_log     = []

    t = threading.Thread(
        target=_stream_script,
        args=("run_demo_scenario.py", ""),
        daemon=True,
    )
    t.start()
    return jsonify({'success': True, 'message': 'Demo scenario started.'})


@app.route('/api/operation-status')
def api_operation_status():
    """Return current operation state and the last N log lines."""
    with ngame_state._operation_lock:
        return jsonify({
            'running':  ngame_state.operation_running,
            'name':     ngame_state.operation_name,
            'log_tail': ngame_state.operation_log[-80:],
        })


# Error Handlers

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

if __name__ == '__main__':
    # Create output directory
    Path("output").mkdir(exist_ok=True)
    Path("uploads").mkdir(exist_ok=True)
    Path("logs").mkdir(exist_ok=True)
    
    # Initialize NGAME state
    if NGAME_AVAILABLE:
        ngame_state.initialize_workflow()
    
    print("🚀 NGAME UI Starting...")
    print(f"📋 NGAME Components Available: {NGAME_AVAILABLE}")
    print(f"👤 FRP mode (hide consultant controls): {FRP_MODE}")
    print("🌐 Open your browser to: http://localhost:5001/dashboard")
    
    # Run the application
    # Disable the auto-reloader to avoid repeated restarts in terminals.
    app.run(debug=True, host='0.0.0.0', port=5001, use_reloader=False)