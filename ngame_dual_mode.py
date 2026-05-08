#!/usr/bin/env python3
"""
NGAME Dual Mode System
Supports both Training Mode and Fraud Analysis Mode with same workflow
but different Phase II tail behaviors.
"""

import os
import json
import numpy as np
from datetime import datetime
from phase1_director_agent_enhanced import EnhancedPhase1DirectorAgent
from phase2_director_agent import Phase2DirectorAgent
from training_agent_excel_manager import TrainingAgentExcelManager

class NGameDualMode:
    """NGAME Dual Mode system supporting both training and fraud analysis."""
    
    def __init__(self):
        self.phase1_agent = EnhancedPhase1DirectorAgent()
        self.phase2_agent = Phase2DirectorAgent()
        self.excel_manager = TrainingAgentExcelManager()
        self.training_matrix_file = "NGAME_Training_Matrix.xlsx"
        self.fraud_analysis_file = "NGAME_Fraud_Analysis.json"
        
    def determine_mode(self):
        """Determine whether to run in Training or Fraud Analysis mode."""
        print("🔍 Determining NGAME Mode")
        print("=" * 40)
        
        # Check if training matrix exists
        training_exists = os.path.exists(self.training_matrix_file)
        
        if not training_exists:
            print("📊 Training Mode: No training matrix found")
            print("🎯 Building churn reference array")
            return "training"
        
        # Load training matrix to check if training period is complete
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(self.training_matrix_file)
            worksheet = workbook.active
            
            # Count training days (excluding Transaction Type and Average columns)
            training_days = worksheet.max_column - 2
            
            if training_days < 30:
                print(f"📊 Training Mode: {training_days} days completed (target: 30+)")
                print("🎯 Continuing to build churn reference array")
                return "training"
            else:
                print(f"🔍 Fraud Analysis Mode: {training_days} days of training data available")
                print("🎯 Analyzing Today's churn for fraud detection")
                return "fraud_analysis"
                
        except Exception as e:
            print(f"❌ Error determining mode: {e}")
            return "training"  # Default to training mode
    
    def execute_phase1(self):
        """Execute Phase I: Extract data and generate Today.ttl."""
        print("📊 Phase I: Data Extraction")
        print("=" * 40)
        
        # Run enhanced Phase 1 with file management
        success = self.phase1_agent.run_enhanced_phase1()
        
        if success:
            print("✅ Phase I completed successfully")
            return True
        else:
            print("❌ Phase I failed")
            return False
    
    def execute_phase2(self):
        """Execute Phase II: Analyze and generate Today's CPI array."""
        print("🔍 Phase II: Fraud Analysis")
        print("=" * 40)
        
        # Check if both TTL files exist
        yesterday_file = 'quickbooks_ontology_Yesterday.ttl'
        today_file = 'quickbooks_ontology_Today.ttl'
        
        if not os.path.exists(yesterday_file):
            print(f"ℹ️  No {yesterday_file} found (first run)")
            print("ℹ️  Phase II will be skipped for first run")
            return None
        
        if not os.path.exists(today_file):
            print(f"❌ {today_file} not found")
            return None
        
        try:
            # Run Phase 2 analysis
            self.phase2_agent.run()
            
            # Extract CPI values from the analysis
            cpi_values = self.extract_cpi_values()
            
            if cpi_values:
                print(f"✅ Phase II completed successfully")
                print(f"📊 Extracted {len(cpi_values)} CPI values")
                return cpi_values
            else:
                print("❌ Failed to extract CPI values")
                return None
                
        except Exception as e:
            print(f"❌ Error in Phase II execution: {e}")
            return None
    
    def extract_cpi_values(self):
        """Extract CPI values from Phase 2 analysis."""
        try:
            # Load the CPI data from Today_Churn_Matrix_Ready.json
            with open('Today_Churn_Matrix_Ready.json', 'r') as f:
                cpi_data = json.load(f)
            
            # Extract similarity measures only
            similarity_values = []
            for label, value in zip(cpi_data['pandas_series_ready']['index'], cpi_data['numpy_array']):
                if "_similarity" in label:
                    similarity_values.append(value)
            
            return similarity_values
            
        except Exception as e:
            print(f"❌ Error extracting CPI values: {e}")
            return None
    
    def execute_training_mode_tail(self, cpi_values):
        """Training Mode Phase II Tail: Add Today's CPI array as next column."""
        print("📈 Training Mode: Adding Today's CPI array to matrix")
        print("=" * 50)
        
        if not cpi_values:
            print("ℹ️  No CPI values to add to training matrix")
            return None
        
        try:
            # Create training matrix if it doesn't exist
            if not os.path.exists(self.training_matrix_file):
                self.create_initial_training_matrix(cpi_values)
                print("✅ Initial training matrix created")
                return 1
            
            # Load existing training matrix
            from openpyxl import load_workbook
            workbook = load_workbook(self.training_matrix_file)
            worksheet = workbook.active
            
            # Get current day number
            current_day = worksheet.max_column - 2  # Subtract 2 for Transaction Type and Average columns
            new_day = current_day + 1
            
            # Add new day header
            new_col = worksheet.max_column + 1
            worksheet.cell(row=1, column=new_col, value=f"Day {new_day}")
            
            # Add new day values
            for i, cpi_value in enumerate(cpi_values, 2):
                worksheet.cell(row=i, column=new_col, value=cpi_value)
                worksheet.cell(row=i, column=new_col).number_format = '0.000000'
            
            # Update μ and σ columns
            self.update_training_averages(worksheet)
            
            # Save updated workbook
            workbook.save(self.training_matrix_file)
            
            print(f"✅ Added Day {new_day} to training matrix")
            print(f"📊 Training matrix now has {new_day} days of data")
            
            return new_day
            
        except Exception as e:
            print(f"❌ Error updating training matrix: {e}")
            return None
    
    def execute_fraud_analysis_mode_tail(self, cpi_values):
        """Fraud Analysis Mode Phase II Tail: Compare and analyze for fraud."""
        print("🔍 Fraud Analysis Mode: Comparing Today's churn with reference")
        print("=" * 60)
        
        if not cpi_values:
            print("ℹ️  No CPI values to analyze")
            return None
        
        try:
            # Load training matrix to get reference averages
            from openpyxl import load_workbook
            workbook = load_workbook(self.training_matrix_file)
            worksheet = workbook.active
            
            # Get reference averages (column B)
            reference_averages = []
            for row in range(2, worksheet.max_row + 1):
                avg_value = worksheet.cell(row=row, column=2).value
                if avg_value is not None:
                    reference_averages.append(avg_value)
            
            # Compare Today's CPI values with reference averages
            comparison_results = self.compare_cpi_arrays(cpi_values, reference_averages)
            
            # Perform fraud analysis
            fraud_analysis = self.perform_fraud_analysis(comparison_results)
            
            # Generate risk assessment
            risk_assessment = self.generate_risk_assessment(fraud_analysis)
            
            # Save fraud analysis results
            fraud_results = {
                "analysis_date": datetime.now().isoformat(),
                "today_cpi_values": cpi_values,
                "reference_averages": reference_averages,
                "comparison_results": comparison_results,
                "fraud_analysis": fraud_analysis,
                "risk_assessment": risk_assessment
            }
            
            with open(self.fraud_analysis_file, 'w') as f:
                json.dump(fraud_results, f, indent=2)
            
            print("✅ Fraud analysis completed")
            print(f"📊 Risk assessment: {risk_assessment['overall_risk']}")
            print(f"📁 Results saved to: {self.fraud_analysis_file}")
            
            return fraud_results
            
        except Exception as e:
            print(f"❌ Error in fraud analysis: {e}")
            return None
    
    def compare_cpi_arrays(self, today_cpi, reference_cpi):
        """Compare Today's CPI values with reference averages."""
        if len(today_cpi) != len(reference_cpi):
            print("⚠️  Warning: CPI arrays have different lengths")
            return None
        
        comparisons = []
        for i, (today_val, ref_val) in enumerate(zip(today_cpi, reference_cpi)):
            deviation = today_val - ref_val
            percent_deviation = (deviation / ref_val) * 100 if ref_val != 0 else 0
            
            comparisons.append({
                "transaction_index": i,
                "today_cpi": today_val,
                "reference_cpi": ref_val,
                "deviation": deviation,
                "percent_deviation": percent_deviation
            })
        
        return comparisons
    
    def perform_fraud_analysis(self, comparison_results):
        """Perform fraud analysis based on comparison results."""
        if not comparison_results:
            return None
        
        # Calculate overall statistics
        deviations = [comp['deviation'] for comp in comparison_results]
        percent_deviations = [comp['percent_deviation'] for comp in comparison_results]
        
        analysis = {
            "total_transactions": len(comparison_results),
            "mean_deviation": np.mean(deviations),
            "std_deviation": np.std(deviations),
            "mean_percent_deviation": np.mean(percent_deviations),
            "high_deviation_count": len([d for d in percent_deviations if abs(d) > 50]),
            "moderate_deviation_count": len([d for d in percent_deviations if 20 <= abs(d) <= 50]),
            "low_deviation_count": len([d for d in percent_deviations if abs(d) < 20])
        }
        
        return analysis
    
    def generate_risk_assessment(self, fraud_analysis):
        """Generate risk assessment based on fraud analysis."""
        if not fraud_analysis:
            return {"overall_risk": "UNKNOWN", "risk_level": "UNKNOWN"}
        
        # Determine overall risk based on deviations
        high_deviation_count = fraud_analysis['high_deviation_count']
        moderate_deviation_count = fraud_analysis['moderate_deviation_count']
        mean_percent_deviation = fraud_analysis['mean_percent_deviation']
        
        if high_deviation_count > 5 or abs(mean_percent_deviation) > 30:
            overall_risk = "HIGH"
            risk_level = "CRITICAL"
        elif high_deviation_count > 2 or abs(mean_percent_deviation) > 15:
            overall_risk = "MEDIUM"
            risk_level = "ELEVATED"
        else:
            overall_risk = "LOW"
            risk_level = "NORMAL"
        
        return {
            "overall_risk": overall_risk,
            "risk_level": risk_level,
            "high_deviation_count": high_deviation_count,
            "moderate_deviation_count": moderate_deviation_count,
            "mean_percent_deviation": mean_percent_deviation
        }
    
    def create_initial_training_matrix(self, cpi_values):
        """Create initial training matrix."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "NGAME Training Matrix"
        
        # Headers
        headers = ["Transaction Type", "μ (Mean CPI)", "σ (Std Dev CPI)", "Day 1"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True, color="FFFFFF")
            cell.fill = cell.fill.copy(start_color="366092", end_color="366092", fill_type="solid")
        
        # Transaction types (assuming standard 17 types)
        transaction_types = [
            "Customers", "Recurring_payments", "Invoices", "Contractors", "Mileage",
            "Products_&_Services", "Chart_of_Accounts", "Sales_orders", "Vendors",
            "Bank_Transactions", "Bill_Payments", "Estimates", "Recurring_Transactions",
            "Bills", "Receipts", "Expense_transactions", "Sales_transactions"
        ]
        
        # Add data rows
        for i, (transaction_type, cpi_value) in enumerate(zip(transaction_types, cpi_values), 2):
            worksheet.cell(row=i, column=1, value=transaction_type)
            worksheet.cell(row=i, column=2, value=cpi_value)
            worksheet.cell(row=i, column=3, value=cpi_value)
            worksheet.cell(row=i, column=2).number_format = '0.000000'
            worksheet.cell(row=i, column=3).number_format = '0.000000'
        
        # Save workbook
        workbook.save(self.training_matrix_file)
    
    def update_training_averages(self, worksheet):
        """Update μ (col B) and σ (col C) in training matrix."""
        # Day columns start at col 4 (A=type, B=μ, C=σ, D+=days)
        day_columns = []
        for col in range(4, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            if header and header.startswith("Day "):
                day_columns.append(col)
        
        # Update average for each row
        for row in range(2, worksheet.max_row + 1):
            # Get all CPI values for this row
            cpi_values = []
            for col in day_columns:
                value = worksheet.cell(row=row, column=col).value
                if value is not None:
                    cpi_values.append(value)
            
            if cpi_values:
                import numpy as _np
                mean_cpi = float(_np.mean(cpi_values))
                std_cpi  = float(_np.std(cpi_values, ddof=1)) if len(cpi_values) >= 2 else 0.0
                worksheet.cell(row=row, column=2, value=mean_cpi)
                worksheet.cell(row=row, column=2).number_format = '0.000000'
                worksheet.cell(row=row, column=3, value=std_cpi)
                worksheet.cell(row=row, column=3).number_format = '0.000000'
    
    def execute_dual_mode(self):
        """Execute NGAME dual mode system."""
        print("🚀 NGAME Dual Mode System")
        print("=" * 60)
        print(f"📅 Execution time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Determine mode
        mode = self.determine_mode()
        print(f"🎯 Selected mode: {mode.upper()}")
        
        # Execute Phase I
        print("\n" + "="*60)
        phase1_success = self.execute_phase1()
        
        if not phase1_success:
            print("❌ Phase I failed - stopping execution")
            return False
        
        # Execute Phase II
        print("\n" + "="*60)
        cpi_values = self.execute_phase2()
        
        # Execute mode-specific Phase II tail
        print("\n" + "="*60)
        if mode == "training":
            result = self.execute_training_mode_tail(cpi_values)
            if result:
                print(f"\n🎉 Training Mode completed successfully!")
                print(f"📈 Training day: {result}")
                print(f"📊 Building churn reference array")
            else:
                print("❌ Training Mode failed")
                return False
        elif mode == "fraud_analysis":
            result = self.execute_fraud_analysis_mode_tail(cpi_values)
            if result:
                print(f"\n🎉 Fraud Analysis Mode completed successfully!")
                print(f"🔍 Risk assessment: {result['risk_assessment']['overall_risk']}")
                print(f"📊 Fraud analysis results generated")
            else:
                print("❌ Fraud Analysis Mode failed")
                return False
        
        return True

def main():
    """Main function for NGAME dual mode execution."""
    print("🚀 NGAME Dual Mode System")
    print("=" * 70)
    
    # Initialize NGAME dual mode
    ngame = NGameDualMode()
    
    # Execute dual mode
    success = ngame.execute_dual_mode()
    
    if success:
        print(f"\n🎉 NGAME execution completed successfully!")
        print(f"📁 Check output files for results")
        return True
    else:
        print(f"\n❌ NGAME execution failed")
        return False

if __name__ == "__main__":
    main()
