#!/usr/bin/env python3
"""
NGAME Sequence CPI Agent  —  Slow-Burn Fraud Detection (SCPI)

Computes the Sequence Churn Proximity Index (SCPI): the geometric mean of
L consecutive daily CPI values per transaction-type position.  Comparing
SCPI against the Training Matrix baseline with a scaled standard error
(σ / √L) amplifies sustained medium-level anomalies into HIGH alerts, even
when no single day crosses the HIGH threshold on its own.

Theory
------
For window L ending at today, per transaction-type k:

    SCPI_k = ( ∏_{i=1}^{L} CPI_k(t-i+1) )^{1/L}     geometric mean of L CPIs

Sequence z-score (same μ/σ as single-day, but divided by √L):

    z_seq_k = ( SCPI_k − μ_k ) / ( σ_k / √L )

The √L factor amplifies: a sustained 2σ/day anomaly becomes:
    z_seq = 2√L  →  for L=5: 4.47 σ  (HIGH)

Key property: No Training Matrix schema changes required.  The existing
μ and σ columns are sufficient.

Deviation bands (same as single-day path):
    |z_seq| > 3.0  →  HIGH
    |z_seq| > 2.0  →  MEDIUM
    else           →  LOW
"""

import json
import math
import os
from datetime import datetime
from typing import Any, Dict, List, Optional
import logging

import numpy as np

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)

from ngame_transaction_types import STANDARD_TRANSACTION_TYPES, CPI_ARRAY_SIZE


class NGameSequenceCpiAgent:
    """
    Reads the rolling CPI buffer written by NGameCpiAnalysisAgent and computes
    the Sequence CPI (SCPI) for a configurable window length L.

    Typical usage inside NGameFraudAnalysisFlowManager::

        seq_agent = NGameSequenceCpiAgent()
        result = seq_agent.compute_scpi(window=5)
        if result['success'] and result['ready']:
            seq_comparison = churn_comparison_agent.compare_sequence_profiles(
                result['scpi_array'],
                training_matrix_file,
                result['window_size'],
                weighted_scpi_array=result.get('weighted_scpi_array'),
            )
    """

    def __init__(self, buffer_file: str = "NGAME_CPI_Rolling_Buffer.json"):
        self.name = "NGameSequenceCpiAgent"
        self.buffer_file = buffer_file
        logger.info(f"🤖 {self.name} initialized  (buffer: {self.buffer_file})")

    # ── Public API ──────────────────────────────────────────────────────────

    def compute_scpi(self, window: int = 5) -> Dict[str, Any]:
        """
        Compute SCPI for the last `window` days and load μ/σ from the
        Training Matrix.  Returns a result dict with keys:

            success        bool
            ready          bool   — False when buffer has fewer than window days
            window_size    int    — L (may be < window when buffer is partial)
            days_available int    — how many days are in the buffer
            dates          list   — ISO date strings for the window used
            scpi_array     list   — 18-element geometric mean CPI
            weighted_scpi_array  list   — 18-element geometric mean wCPI (or None)
            scpi_z_scores        list   — 18-element sequence z-scores
            deviation_levels     list   — 18-element ["LOW"|"MEDIUM"|"HIGH"]
            high_count     int
            medium_count   int
            scpi_summary   dict   — mean/std/min/max of scpi_array
            computation_timestamp  str
        """
        logger.info(f"🔍 {self.name}: Computing SCPI (window={window})")

        try:
            buffer = self._load_buffer()
            days_available = len(buffer)

            if days_available == 0:
                logger.info(f"ℹ️  {self.name}: Buffer empty — sequence path skipped")
                return self._not_ready(window, 0)

            if days_available < window:
                logger.info(
                    f"ℹ️  {self.name}: Buffer has {days_available}/{window} days "
                    f"— sequence path not yet active"
                )
                return self._not_ready(window, days_available)

            window_records = buffer[-window:]
            dates = [r.get('date', '') for r in window_records]

            # Validate all records have the expected CPI array size
            for r in window_records:
                if len(r.get('cpi_array', [])) != CPI_ARRAY_SIZE:
                    logger.warning(
                        f"⚠️  {self.name}: Record for {r.get('date')} has "
                        f"cpi_array length {len(r.get('cpi_array', []))}, "
                        f"expected {CPI_ARRAY_SIZE} — skipping sequence"
                    )
                    return self._not_ready(window, days_available)

            cpi_matrix = [r['cpi_array'] for r in window_records]
            scpi_array = self._geometric_mean_columns(cpi_matrix)

            # Dollar-weighted SCPI (only if all records carry a weighted array)
            weighted_scpi_array: Optional[List[float]] = None
            if all(len(r.get('weighted_cpi_array', [])) == CPI_ARRAY_SIZE
                   for r in window_records):
                w_matrix = [r['weighted_cpi_array'] for r in window_records]
                weighted_scpi_array = self._geometric_mean_columns(w_matrix)

            # Sequence z-scores require μ and σ from the Training Matrix
            z_scores, deviation_levels = self._compute_z_scores(
                scpi_array, window
            )

            high_count   = deviation_levels.count('HIGH')
            medium_count = deviation_levels.count('MEDIUM')

            scpi_summary = self._array_summary(scpi_array)

            result: Dict[str, Any] = {
                'success':         True,
                'ready':           True,
                'window_size':     window,
                'days_available':  days_available,
                'dates':           dates,
                'scpi_array':      scpi_array,
                'weighted_scpi_array': weighted_scpi_array,
                'scpi_z_scores':   z_scores,
                'deviation_levels': deviation_levels,
                'high_count':      high_count,
                'medium_count':    medium_count,
                'scpi_summary':    scpi_summary,
                'computation_timestamp': datetime.now().isoformat(),
            }

            logger.info(
                f"✅ {self.name}: SCPI computed over {window} days  "
                f"HIGH={high_count}  MEDIUM={medium_count}"
            )
            if high_count > 0:
                idxs = [i for i, d in enumerate(deviation_levels) if d == 'HIGH']
                logger.warning(
                    f"🚨 {self.name}: Slow-burn HIGH signal on positions: "
                    f"{[STANDARD_TRANSACTION_TYPES[i] for i in idxs]}"
                )

            return result

        except Exception as e:
            logger.error(f"❌ {self.name}: compute_scpi failed — {e}")
            return {'success': False, 'ready': False, 'error': str(e)}

    def get_buffer_summary(self) -> Dict[str, Any]:
        """Return a summary of the current rolling buffer (useful for dashboards)."""
        buffer = self._load_buffer()
        return {
            'days_stored': len(buffer),
            'oldest_date': buffer[0].get('date')  if buffer else None,
            'newest_date': buffer[-1].get('date') if buffer else None,
            'buffer_file': self.buffer_file,
        }

    # ── Internal helpers ────────────────────────────────────────────────────

    def _load_buffer(self) -> List[Dict]:
        if not os.path.exists(self.buffer_file):
            return []
        try:
            with open(self.buffer_file, 'r') as fh:
                data = json.load(fh)
            return data if isinstance(data, list) else []
        except Exception as e:
            logger.error(f"❌ {self.name}: Cannot read buffer file — {e}")
            return []

    @staticmethod
    def _geometric_mean_columns(matrix: List[List[float]]) -> List[float]:
        """
        Compute the per-column geometric mean of a 2-D list (L rows × 18 cols).

        Computed in log space for numerical stability:
            geo_mean_k = exp( mean( log(max(CPI_k, ε)) ) )

        ε = 1e-9 guards against log(0) when CPI is exactly zero (complete
        turnover day) — an extreme but possible edge case.
        """
        eps = 1e-9
        arr = np.array(matrix, dtype=float)
        arr = np.clip(arr, eps, None)
        return list(np.exp(np.mean(np.log(arr), axis=0)))

    def _compute_z_scores(
        self,
        scpi_array: List[float],
        window: int,
        training_matrix_file: str = "NGAME_Training_Matrix.xlsx",
    ) -> tuple:
        """
        Load μ/σ from the Training Matrix and compute per-position sequence z-scores.

        z_seq_k = (SCPI_k − μ_k) / (σ_k / √L)

        Falls back to z_seq = 0.0 (LOW) when σ is zero or the Training Matrix
        is unavailable — same defensive behaviour as NGameChurnComparisonAgent.
        """
        mu    = self._load_column(training_matrix_file, col=2)  # column B
        sigma = self._load_column(training_matrix_file, col=3)  # column C

        sqrt_L = math.sqrt(window)
        z_scores: List[float]      = []
        deviation_levels: List[str] = []

        for k in range(CPI_ARRAY_SIZE):
            mu_k    = mu[k]    if k < len(mu)    else 0.0
            sigma_k = sigma[k] if k < len(sigma) else 0.0

            if sigma_k > 0:
                z = (scpi_array[k] - mu_k) / (sigma_k / sqrt_L)
            else:
                z = 0.0

            z_scores.append(z)

            abs_z = abs(z)
            if abs_z > 3.0:
                deviation_levels.append('HIGH')
            elif abs_z > 2.0:
                deviation_levels.append('MEDIUM')
            else:
                deviation_levels.append('LOW')

        return z_scores, deviation_levels

    @staticmethod
    def _load_column(
        training_matrix_file: str, col: int
    ) -> List[float]:
        """Load a single column (1-indexed) from the Training Matrix Excel file."""
        if not os.path.exists(training_matrix_file):
            logger.warning(
                f"⚠️  NGameSequenceCpiAgent: Training Matrix not found "
                f"({training_matrix_file}) — z-scores will be 0"
            )
            return [0.0] * CPI_ARRAY_SIZE
        try:
            from openpyxl import load_workbook
            wb = load_workbook(training_matrix_file)
            ws = wb.active
            values: List[float] = []
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col).value
                try:
                    values.append(float(cell) if cell is not None else 0.0)
                except (ValueError, TypeError):
                    values.append(0.0)
            return values
        except Exception as e:
            logger.warning(f"⚠️  NGameSequenceCpiAgent: Cannot load column {col} — {e}")
            return [0.0] * CPI_ARRAY_SIZE

    @staticmethod
    def _array_summary(arr: List[float]) -> Dict[str, float]:
        a = np.array(arr)
        return {
            'mean':  float(np.mean(a)),
            'std':   float(np.std(a)),
            'min':   float(np.min(a)),
            'max':   float(np.max(a)),
        }

    @staticmethod
    def _not_ready(window: int, days_available: int) -> Dict[str, Any]:
        return {
            'success':        True,
            'ready':          False,
            'window_size':    window,
            'days_available': days_available,
        }


def main():
    """Quick standalone test of the SCPI agent."""
    print("🚀 NGAME Sequence CPI Agent Test")
    print("=" * 50)

    agent = NGameSequenceCpiAgent()

    summary = agent.get_buffer_summary()
    print(f"\n📊 Rolling Buffer:  {summary['days_stored']} day(s) stored")
    if summary['oldest_date']:
        print(f"   Range: {summary['oldest_date']}  →  {summary['newest_date']}")

    for L in (3, 5, 7):
        result = agent.compute_scpi(window=L)
        if not result['success']:
            print(f"\n❌ window={L}: {result.get('error')}")
        elif not result['ready']:
            print(
                f"\n⏳ window={L}: not yet active "
                f"({result['days_available']}/{L} days in buffer)"
            )
        else:
            print(f"\n✅ window={L}  ({', '.join(result['dates'])})")
            print(
                f"   SCPI mean={result['scpi_summary']['mean']:.6f}  "
                f"std={result['scpi_summary']['std']:.6f}"
            )
            print(
                f"   HIGH={result['high_count']}  "
                f"MEDIUM={result['medium_count']}"
            )
            for i, (tx, z, lvl) in enumerate(zip(
                STANDARD_TRANSACTION_TYPES,
                result['scpi_z_scores'],
                result['deviation_levels'],
            )):
                if lvl != 'LOW':
                    print(f"   ⚠️  φ{i+1:2d} {tx:<30s}  z_seq={z:+.3f}  {lvl}")


if __name__ == "__main__":
    main()
