#!/usr/bin/env python3
"""
Merge Day columns from a replay matrix into the completed 30-day matrix, then
recalculate μ (column B) and σ (column C) from all Day columns.

Typical use after replaying Sessions 1–12 with the fixed type-alias code:

  python3 merge_training_matrix_days.py \\
    --target NGAME_Training_Matrix.xlsx \\
    --source NGAME_Training_Matrix_REPLAY_D12.xlsx \\
    --days 1-12

Also merges matching wDay columns on the wCPI_Baseline sheet (wDay 2+ only;
Day 1 has no dollar-weighted column because there is no Yesterday snapshot).

Always back up the target matrix first. This script creates a timestamped
.backup copy before writing unless --no-backup is passed.
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Tuple

import numpy as np
from openpyxl import load_workbook


def _parse_day_range(spec: str) -> List[int]:
    """Parse '1-11', '1,2,3', or '11'."""
    days: List[int] = []
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start_s, end_s = part.split("-", 1)
            start, end = int(start_s), int(end_s)
            if start > end:
                raise ValueError(f"Invalid range: {part}")
            days.extend(range(start, end + 1))
        else:
            days.append(int(part))
    if not days:
        raise ValueError("No day numbers parsed")
    return sorted(set(days))


def _day_columns(worksheet) -> List[Tuple[int, int]]:
    """Return (day_number, column_index) for each Day N header."""
    cols: List[Tuple[int, int]] = []
    for col in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1, column=col).value
        if not header:
            continue
        m = re.match(r"^Day\s+(\d+)$", str(header).strip(), re.I)
        if m:
            cols.append((int(m.group(1)), col))
    return cols


def _find_row_for_type(worksheet, tx_type: str) -> int | None:
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == tx_type:
            return row
    return None


def recalc_mu_sigma(worksheet) -> None:
    """Recalculate μ (col B) and σ (col C) from all Day columns (col D+)."""
    day_cols = [col for _, col in _day_columns(worksheet)]
    if not day_cols:
        raise ValueError("No Day columns found in worksheet")

    for row in range(2, worksheet.max_row + 1):
        values: List[float] = []
        for col in day_cols:
            val = worksheet.cell(row=row, column=col).value
            if val is None:
                continue
            try:
                values.append(float(val))
            except (TypeError, ValueError):
                pass

        if len(values) >= 1:
            mu = float(np.mean(values))
            sigma = float(np.std(values, ddof=1)) if len(values) >= 2 else 0.0
        else:
            mu, sigma = 0.0, 0.0

        worksheet.cell(row=row, column=2, value=mu)
        worksheet.cell(row=row, column=3, value=sigma)


def _wday_columns(worksheet) -> List[Tuple[int, int]]:
    """Return (day_number, column_index) for each wDay N header."""
    cols: List[Tuple[int, int]] = []
    for col in range(1, worksheet.max_column + 1):
        header = worksheet.cell(row=1, column=col).value
        if not header:
            continue
        m = re.match(r"^wDay\s+(\d+)$", str(header).strip(), re.I)
        if m:
            cols.append((int(m.group(1)), col))
    return cols


def recalc_w_mu_sigma(worksheet) -> None:
    """Recalculate μ_w (col B) and σ_w (col C) from all wDay columns."""
    wday_cols = [col for _, col in _wday_columns(worksheet)]
    if not wday_cols:
        return

    for row in range(2, worksheet.max_row + 1):
        values: List[float] = []
        for col in wday_cols:
            val = worksheet.cell(row=row, column=col).value
            if val is None:
                continue
            try:
                values.append(float(val))
            except (TypeError, ValueError):
                pass

        if len(values) >= 1:
            mu = float(np.mean(values))
            sigma = float(np.std(values, ddof=1)) if len(values) >= 2 else 0.0
        else:
            mu, sigma = 0.0, 0.0

        worksheet.cell(row=row, column=2, value=mu)
        worksheet.cell(row=row, column=3, value=sigma)


def _merge_sheet_days(
    tgt_ws,
    src_ws,
    days: List[int],
    day_map_fn,
    header_prefix: str,
) -> int:
    tgt_map = dict(day_map_fn(tgt_ws))
    src_map = dict(day_map_fn(src_ws))

    missing_src = [d for d in days if d not in src_map]
    if missing_src:
        raise ValueError(f"Source missing {header_prefix} column(s): {missing_src}")
    missing_tgt = [d for d in days if d not in tgt_map]
    if missing_tgt:
        raise ValueError(f"Target missing {header_prefix} column(s): {missing_tgt}")

    copied = 0
    for day in days:
        src_col = src_map[day]
        tgt_col = tgt_map[day]
        for row in range(2, min(tgt_ws.max_row, src_ws.max_row) + 1):
            tx_type = tgt_ws.cell(row=row, column=1).value
            if tx_type is None:
                continue
            src_row = _find_row_for_type(src_ws, str(tx_type))
            if src_row is None:
                print(f"  Warning: {tx_type!r} not in source — skip {header_prefix} {day}")
                continue
            val = src_ws.cell(row=src_row, column=src_col).value
            tgt_ws.cell(row=row, column=tgt_col, value=val)
            copied += 1
    return copied


def merge_days(
    target_path: Path,
    source_path: Path,
    days: List[int],
    no_backup: bool,
) -> None:
    if not target_path.exists():
        raise FileNotFoundError(f"Target matrix not found: {target_path}")
    if not source_path.exists():
        raise FileNotFoundError(f"Source matrix not found: {source_path}")

    if not no_backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = target_path.with_name(f"{target_path.stem}_pre_merge_{stamp}{target_path.suffix}")
        shutil.copy2(target_path, backup)
        print(f"Backup: {backup}")

    tgt_wb = load_workbook(target_path)
    src_wb = load_workbook(source_path, read_only=True)

    sheet_name = "Training Matrix"
    tgt_ws = tgt_wb[sheet_name] if sheet_name in tgt_wb.sheetnames else tgt_wb.active
    src_ws = src_wb[sheet_name] if sheet_name in src_wb.sheetnames else src_wb.active

    _merge_sheet_days(tgt_ws, src_ws, days, _day_columns, "Day")
    print(f"Merged Day(s) {days} on Training Matrix from {source_path.name}")

    recalc_mu_sigma(tgt_ws)
    print("Recalculated μ (col B) and σ (col C) from all Day columns.")

    wcpi_name = "wCPI_Baseline"
    wcpi_days = [d for d in days if d >= 2]
    if wcpi_days and wcpi_name in tgt_wb.sheetnames and wcpi_name in src_wb.sheetnames:
        _merge_sheet_days(
            tgt_wb[wcpi_name], src_wb[wcpi_name], wcpi_days, _wday_columns, "wDay"
        )
        print(f"Merged wDay(s) {wcpi_days} on {wcpi_name} sheet.")
        recalc_w_mu_sigma(tgt_wb[wcpi_name])
        print("Recalculated μ_w (col B) and σ_w (col C) on wCPI_Baseline.")
    elif wcpi_days:
        print(f"Note: {wcpi_name} sheet missing in source or target — skipped wCPI merge.")

    src_wb.close()

    tgt_wb.save(target_path)
    print(f"Saved: {target_path}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge replay Day columns into 30-day matrix.")
    parser.add_argument(
        "--target",
        default="NGAME_Training_Matrix.xlsx",
        help="Completed 30-day matrix to update (default: NGAME_Training_Matrix.xlsx)",
    )
    parser.add_argument(
        "--source",
        required=True,
        help="Replay matrix containing corrected Day columns (e.g. NGAME_Training_Matrix_REPLAY_D11.xlsx)",
    )
    parser.add_argument(
        "--days",
        default="1-12",
        help="Day numbers to copy, e.g. '1-12' or '1,2,3' (default: 1-12)",
    )
    parser.add_argument("--no-backup", action="store_true", help="Do not create a pre-merge backup")
    args = parser.parse_args()

    days = _parse_day_range(args.days)
    try:
        merge_days(Path(args.target), Path(args.source), days, args.no_backup)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
