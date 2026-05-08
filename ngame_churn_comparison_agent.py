#!/usr/bin/env python3
"""
NGAME Churn Comparison Agent
Handles comparison of Today's churn profile with stored Average CPI array from Training Matrix
"""

import os
import json
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class NGameChurnComparisonAgent:
    """
    NGAME Churn Comparison Agent for fraud analysis flow.
    Handles comparison of Today's churn profile with stored reference array.
    """
    
    def __init__(self):
        self.name = "NGameChurnComparisonAgent"
        self.cpi_array_size = 18  # 18 transaction types (incl. Vendors φ18)
        
        logger.info(f"🤖 {self.name} initialized")
    
    def compare_churn_profiles(
        self,
        today_cpi_array: List[float],
        training_matrix_file: str,
        weighted_cpi_array: Optional[List[float]] = None,
    ) -> Dict[str, Any]:
        """
        Compare Today's churn profile with stored μ and σ from the Training Matrix.

        Primary signal  — z-score against training baseline:
            |z| > 3.0  → HIGH    (>3σ from mean)
            |z| > 2.0  → MEDIUM  (>2σ from mean)
            else       → LOW

        Secondary signal (when weighted_cpi_array is supplied) — dollar-weight divergence:
            dollar_weight_divergence = |weighted_cpi − unweighted_cpi|  for each position.
            A large divergence means a small number of high-value transactions are driving
            the day's churn — the "$50,000 vs $50" alarm.

            Divergence bands (tunable):
            > 0.30  → dollar_alarm = HIGH
            > 0.15  → dollar_alarm = MEDIUM
            else    → dollar_alarm = LOW

        composite_alarm fires when EITHER signal is HIGH or MEDIUM.
        Implements Requirement 3 from Chart 2.
        """
        logger.info(f"🔍 {self.name}: Comparing churn profiles")
        
        try:
            # Load μ (col B) and σ (col C) from training matrix
            average_cpi_array = self._load_average_cpi_array(training_matrix_file)
            std_cpi_array     = self._load_std_cpi_array(training_matrix_file)
            
            if not average_cpi_array:
                return {
                    'success': False,
                    'error': 'Failed to load μ array from training matrix'
                }
            
            # Validate array sizes
            if len(today_cpi_array) != self.cpi_array_size or len(average_cpi_array) != self.cpi_array_size:
                return {
                    'success': False,
                    'error': f'Array size mismatch: Today={len(today_cpi_array)}, Average={len(average_cpi_array)}, Expected={self.cpi_array_size}'
                }
            
            # ── Dollar-weighted baseline (trained μ_w / σ_w) ─────────────────────
            # Attempt to load the trained dollar-weighted baseline from the second
            # sheet in the Training Matrix.  Falls back to the divergence-threshold
            # signal when fewer than 2 wCPI training days have been recorded.
            dollar_baseline = self._load_dollar_weighted_baseline(training_matrix_file)

            # Calculate z-score differences for each element pair
            differences = self._calculate_differences(today_cpi_array, average_cpi_array, std_cpi_array)
            
            # Calculate overall array distance metrics
            distance_metrics = self._calculate_array_distance_metrics(today_cpi_array, average_cpi_array)
            
            # Calculate summary statistics
            summary_stats = self._calculate_summary_statistics(differences, distance_metrics)
            
            # Convert arrays to 1-17 indexed dictionaries
            today_cpi_dict   = {str(i + 1): value for i, value in enumerate(today_cpi_array)}
            average_cpi_dict = {str(i + 1): value for i, value in enumerate(average_cpi_array)}
            std_cpi_dict     = {str(i + 1): value for i, value in enumerate(std_cpi_array)}
            
            # ── Dollar-weight signal ──────────────────────────────────────────────
            has_dollar_weighting = (
                weighted_cpi_array is not None
                and len(weighted_cpi_array) == len(today_cpi_array)
            )
            if has_dollar_weighting:
                if dollar_baseline['available']:
                    # PRIMARY path: z_dollar = (wCPI − μ_w) / σ_w
                    dollar_signals = self._calculate_dollar_z_score_signal(
                        weighted_cpi_array,
                        dollar_baseline['mu_w'],
                        dollar_baseline['sigma_w'],
                    )
                    dollar_method = 'z_score'
                else:
                    # FALLBACK path (pre-training): raw |wCPI − CPI| divergence
                    dollar_signals = self._calculate_dollar_weight_signal(
                        today_cpi_array, weighted_cpi_array
                    )
                    dollar_method = 'divergence_fallback'

                dollar_alarm_count = 0
                for i, diff in enumerate(differences):
                    sig = dollar_signals[i] if i < len(dollar_signals) else {}
                    diff['weighted_cpi']             = weighted_cpi_array[i]
                    diff['dollar_alarm_level']        = sig.get('dollar_level', 'LOW')
                    diff['dollar_alarm']              = sig.get('dollar_alarm', False)
                    diff['dollar_z_score']            = sig.get('dollar_z_score', 0.0)
                    diff['dollar_abs_z_score']        = sig.get('dollar_abs_z_score', 0.0)
                    diff['dollar_weight_divergence']  = sig.get('divergence', 0.0)
                    diff['dollar_signal_method']      = dollar_method
                    # Composite: fire when EITHER count-based OR dollar z-score is elevated
                    diff['composite_alarm'] = (
                        diff['deviation_level'] in ('HIGH', 'MEDIUM')
                        or diff['dollar_alarm']
                    )
                    if diff['dollar_alarm']:
                        dollar_alarm_count += 1

                logger.info(
                    f"💵 Dollar-weight signal ({dollar_method}): "
                    f"{dollar_alarm_count} position(s) elevated"
                )
            else:
                for diff in differences:
                    diff['composite_alarm'] = diff['deviation_level'] in ('HIGH', 'MEDIUM')

            result = {
                'success': True,
                'today_cpi_array':          today_cpi_dict,
                'average_cpi_array':        average_cpi_dict,
                'std_cpi_array':            std_cpi_dict,
                'differences':              differences,
                'distance_metrics':         distance_metrics,
                'summary_stats':            summary_stats,
                'has_dollar_weighting':     has_dollar_weighting,
                'dollar_baseline_available': dollar_baseline['available'],
                'dollar_baseline_days':     dollar_baseline.get('training_days_w', 0),
                'comparison_timestamp':     datetime.now().isoformat(),
            }
            if has_dollar_weighting:
                result['weighted_cpi_array'] = {
                    str(i + 1): v for i, v in enumerate(weighted_cpi_array)
                }

            logger.info(f"✅ {self.name}: Churn profile comparison completed successfully")
            logger.info(f"📊 Mean |z-score|: {summary_stats['mean_abs_z_score']:.4f}")
            logger.info(f"📊 High-deviation count (|z|>3): {summary_stats['high_deviation_count']}")
            logger.info(f"📏 Euclidean distance: {distance_metrics['euclidean_distance']:.6f}")
            logger.info(f"📏 Manhattan distance: {distance_metrics['manhattan_distance']:.6f}")

            return result
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Churn profile comparison failed - {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _load_average_cpi_array(self, training_matrix_file: str) -> Optional[List[float]]:
        """Load μ (Average CPI) array from column B of the training matrix Excel file."""
        logger.info(f"📊 {self.name}: Loading μ array from {training_matrix_file}")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(training_matrix_file)
            worksheet = workbook.active
            
            average_cpi_array = []
            for row in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=2).value
                if cell_value is not None:
                    try:
                        average_cpi_array.append(float(cell_value))
                    except (ValueError, TypeError):
                        logger.warning(f"⚠️  Invalid μ value in row {row}, col B: {cell_value}")
                        average_cpi_array.append(0.0)
                else:
                    average_cpi_array.append(0.0)
            
            logger.info(f"✅ {self.name}: Loaded μ array with {len(average_cpi_array)} elements")
            return average_cpi_array
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Error loading μ array: {e}")
            return None

    def _load_std_cpi_array(self, training_matrix_file: str) -> List[float]:
        """
        Load σ (Std Dev CPI) array from column C of the training matrix Excel file.
        Returns a list of zeros if the column is absent (pre-σ matrix or first-day matrix).
        """
        logger.info(f"📊 {self.name}: Loading σ array from {training_matrix_file}")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(training_matrix_file)
            worksheet = workbook.active
            
            std_cpi_array = []
            for row in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=3).value
                if cell_value is not None:
                    try:
                        std_cpi_array.append(float(cell_value))
                    except (ValueError, TypeError):
                        logger.warning(f"⚠️  Invalid σ value in row {row}, col C: {cell_value}")
                        std_cpi_array.append(0.0)
                else:
                    std_cpi_array.append(0.0)
            
            logger.info(f"✅ {self.name}: Loaded σ array with {len(std_cpi_array)} elements")
            return std_cpi_array
            
        except Exception as e:
            logger.warning(f"⚠️  {self.name}: Could not load σ array ({e}); defaulting to zeros")
            return [0.0] * self.cpi_array_size
    
    def _calculate_differences(
        self,
        today_array: List[float],
        average_array: List[float],
        std_array: List[float]
    ) -> List[Dict[str, Any]]:
        """
        Calculate per-element z-score differences between today's CPI and training μ/σ.

        z_score = (today − μ) / σ   (0.0 when σ = 0 to avoid division by zero)
        Deviation bands: |z| > 3.0 → HIGH, |z| > 2.0 → MEDIUM, else LOW.

        Fraud emphasis: a large *negative* z-score (low CPI) signals mass deletions;
        a large *positive* z-score signals a bulk-insertion day.
        """
        logger.info(f"🔍 {self.name}: Calculating z-score differences for {len(today_array)} elements")
        
        differences = []
        
        for i in range(len(today_array)):
            today_value   = today_array[i]
            mean_value    = average_array[i]
            std_value     = std_array[i] if i < len(std_array) else 0.0
            
            absolute_difference = abs(today_value - mean_value)
            
            # z-score (signed — negative means lower CPI than baseline, i.e. more churn)
            if std_value > 0:
                z_score = (today_value - mean_value) / std_value
            else:
                z_score = 0.0
            
            abs_z = abs(z_score)
            if abs_z > 3.0:
                deviation_level = "HIGH"
            elif abs_z > 2.0:
                deviation_level = "MEDIUM"
            else:
                deviation_level = "LOW"
            
            # Retain percent_deviation as a secondary diagnostic metric
            if mean_value != 0:
                percent_deviation = (absolute_difference / abs(mean_value)) * 100
            else:
                percent_deviation = 100.0 if today_value != 0 else 0.0
            
            differences.append({
                'index':               i + 1,   # 1-17 numbering
                'today_value':         today_value,
                'average_value':       mean_value,
                'std_value':           std_value,
                'absolute_difference': absolute_difference,
                'z_score':             z_score,
                'abs_z_score':         abs_z,
                'percent_deviation':   percent_deviation,
                'deviation_level':     deviation_level,
            })
        
        logger.info(f"✅ {self.name}: Calculated z-score differences for {len(differences)} elements")
        return differences
    
    def _calculate_array_distance_metrics(self, today_array: List[float], average_array: List[float]) -> Dict[str, float]:
        """
        Calculate overall array-to-array distance metrics.
        Computes Euclidean, Manhattan, and Cosine distance between the two arrays.
        """
        logger.info(f"📏 {self.name}: Calculating array distance metrics")
        
        try:
            # Convert to numpy arrays for efficient computation
            today_np = np.array(today_array)
            average_np = np.array(average_array)
            
            # Euclidean Distance: √(Σ(today[i] - average[i])²)
            euclidean_distance = float(np.linalg.norm(today_np - average_np))
            
            # Manhattan Distance: Σ|today[i] - average[i]|
            manhattan_distance = float(np.sum(np.abs(today_np - average_np)))
            
            # Cosine Distance: 1 - cosine_similarity
            # Cosine similarity = (today · average) / (||today|| × ||average||)
            dot_product = np.dot(today_np, average_np)
            norm_today = np.linalg.norm(today_np)
            norm_average = np.linalg.norm(average_np)
            
            if norm_today == 0 or norm_average == 0:
                cosine_similarity = 0.0
                cosine_distance = 1.0
            else:
                cosine_similarity = dot_product / (norm_today * norm_average)
                cosine_distance = 1.0 - cosine_similarity
            
            # Mean Squared Error (MSE): (1/n) × Σ(today[i] - average[i])²
            mse = float(np.mean((today_np - average_np) ** 2))
            
            # Root Mean Squared Error (RMSE): √MSE
            rmse = float(np.sqrt(mse))
            
            distance_metrics = {
                'euclidean_distance': euclidean_distance,
                'manhattan_distance': manhattan_distance,
                'cosine_distance': cosine_distance,
                'cosine_similarity': float(cosine_similarity),
                'mean_squared_error': mse,
                'root_mean_squared_error': rmse
            }
            
            logger.info(f"✅ {self.name}: Array distance metrics calculated")
            logger.info(f"📏 Euclidean distance: {euclidean_distance:.6f}")
            logger.info(f"📏 Manhattan distance: {manhattan_distance:.6f}")
            logger.info(f"📏 Cosine distance: {cosine_distance:.6f}")
            
            return distance_metrics
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Error calculating distance metrics: {e}")
            return {
                'euclidean_distance': 0.0,
                'manhattan_distance': 0.0,
                'cosine_distance': 0.0,
                'cosine_similarity': 0.0,
                'mean_squared_error': 0.0,
                'root_mean_squared_error': 0.0
            }
    
    def _calculate_summary_statistics(self, differences: List[Dict[str, Any]], distance_metrics: Dict[str, float]) -> Dict[str, Any]:
        """Calculate summary statistics for differences (z-score based)."""
        logger.info(f"📊 {self.name}: Calculating summary statistics")
        
        absolute_differences = [diff['absolute_difference'] for diff in differences]
        z_scores             = [diff['z_score']             for diff in differences]
        abs_z_scores         = [diff['abs_z_score']         for diff in differences]
        percent_deviations   = [diff['percent_deviation']   for diff in differences]
        deviation_levels     = [diff['deviation_level']     for diff in differences]
        
        high_count   = deviation_levels.count("HIGH")
        medium_count = deviation_levels.count("MEDIUM")
        low_count    = deviation_levels.count("LOW")
        
        summary_stats = {
            # Z-score statistics (primary)
            'mean_abs_z_score':  float(np.mean(abs_z_scores)),
            'max_abs_z_score':   float(np.max(abs_z_scores)),
            'min_abs_z_score':   float(np.min(abs_z_scores)),
            'mean_z_score':      float(np.mean(z_scores)),
            'std_z_score':       float(np.std(z_scores)),
            # Absolute difference statistics (secondary)
            'mean_difference':   float(np.mean(absolute_differences)),
            'std_difference':    float(np.std(absolute_differences)),
            'max_difference':    float(np.max(absolute_differences)),
            'min_difference':    float(np.min(absolute_differences)),
            # Percent deviation statistics (secondary, diagnostic only)
            'mean_percent_deviation': float(np.mean(percent_deviations)),
            'max_percent_deviation':  float(np.max(percent_deviations)),
            # Deviation level counts
            'high_deviation_count':   high_count,
            'medium_deviation_count': medium_count,
            'low_deviation_count':    low_count,
            'total_elements':         len(differences),
            # Array-to-array distance metrics
            'euclidean_distance':      distance_metrics['euclidean_distance'],
            'manhattan_distance':      distance_metrics['manhattan_distance'],
            'cosine_distance':         distance_metrics['cosine_distance'],
            'cosine_similarity':       distance_metrics['cosine_similarity'],
            'mean_squared_error':      distance_metrics['mean_squared_error'],
            'root_mean_squared_error': distance_metrics['root_mean_squared_error'],
        }
        
        logger.info(f"✅ {self.name}: Summary statistics calculated")
        return summary_stats
    
    def _calculate_dollar_weight_signal(
        self,
        unweighted_cpi: List[float],
        weighted_cpi: List[float],
    ) -> List[Dict[str, Any]]:
        """
        Compute the per-position dollar-weight divergence signal.

        dollar_weight_divergence = |weighted_cpi[i] − unweighted_cpi[i]|

        A divergence close to zero means today's new transactions are proportional
        in dollar value to the retained ones — normal.  A large divergence means
        a few high-value items are distorting the count-based view — alarm.

        Divergence thresholds:
            > 0.30  → HIGH   (dollar amounts grossly disproportionate to counts)
            > 0.15  → MEDIUM (noticeable dollar imbalance worth reviewing)
            ≤ 0.15  → LOW

        Direction:
            weighted < unweighted  →  'heavy_additions'  (new items have outsized value)
            weighted > unweighted  →  'heavy_retentions' (retained items dominate in value)
        """
        signal = []
        for uw, w in zip(unweighted_cpi, weighted_cpi):
            divergence = abs(w - uw)
            if divergence > 0.30:
                dollar_alarm = True
                dollar_level = 'HIGH'
            elif divergence > 0.15:
                dollar_alarm = True
                dollar_level = 'MEDIUM'
            else:
                dollar_alarm = False
                dollar_level = 'LOW'

            direction = 'heavy_additions' if w < uw else 'heavy_retentions'

            signal.append({
                'divergence':   divergence,
                'direction':    direction,
                'dollar_level': dollar_level,
                'dollar_alarm': dollar_alarm,
            })
        return signal

    def _calculate_dollar_z_score_signal(
        self,
        weighted_cpi: List[float],
        mu_w: List[float],
        sigma_w: List[float],
    ) -> List[Dict[str, Any]]:
        """
        Compute the per-position dollar z-score signal using the trained baseline.

        z_dollar_k = (wCPI_k − μ_w_k) / σ_w_k

        Same deviation bands as the count-based z-score:
            |z_dollar| > 3.0  →  HIGH
            |z_dollar| > 2.0  →  MEDIUM
            else              →  LOW

        This is the PRIMARY dollar signal once at least 2 wCPI training days
        have been collected.  It correctly accounts for the fact that some
        transaction types (e.g. Bill_Payments) normally carry large dollar
        amounts — a high wCPI divergence for those types is expected and should
        NOT fire an alarm unless it deviates from its own trained baseline.
        """
        signal = []
        for k, (w, mu, sigma) in enumerate(zip(weighted_cpi, mu_w, sigma_w)):
            if sigma > 0:
                z = (w - mu) / sigma
            else:
                z = 0.0

            abs_z = abs(z)
            if abs_z > 3.0:
                dollar_alarm = True
                dollar_level = 'HIGH'
            elif abs_z > 2.0:
                dollar_alarm = True
                dollar_level = 'MEDIUM'
            else:
                dollar_alarm = False
                dollar_level = 'LOW'

            signal.append({
                'dollar_z_score':    z,
                'dollar_abs_z_score': abs_z,
                'dollar_level':      dollar_level,
                'dollar_alarm':      dollar_alarm,
                'divergence':        0.0,   # not used in this path
            })
        return signal

    def _load_dollar_weighted_baseline(
        self, training_matrix_file: str
    ) -> Dict[str, Any]:
        """
        Load μ_w and σ_w from the 'wCPI_Baseline' sheet of the Training Matrix.

        Delegates to NGameMatrixManagementAgent to keep all sheet-access logic
        in one place.  Returns a safe default when the sheet doesn't exist.
        """
        zeros = [0.0] * self.cpi_array_size
        try:
            from ngame_matrix_management_agent import NGameMatrixManagementAgent
            mgr = NGameMatrixManagementAgent()
            # Override the matrix file path in case it differs from the default
            mgr.matrix_file = training_matrix_file
            return mgr.load_dollar_weighted_baseline()
        except Exception as e:
            logger.warning(
                f"⚠️  {self.name}: Cannot load dollar-weighted baseline "
                f"({e}) — falling back to divergence signal"
            )
            return {'available': False, 'mu_w': zeros, 'sigma_w': zeros,
                    'training_days_w': 0}

    def rank_differences(self, differences: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Rank differences by absolute z-score (highest anomaly first).
        Implements Requirement 4 from Chart 2.
        """
        logger.info(f"📊 {self.name}: Ranking differences by |z-score|")
        
        try:
            ranked_differences = sorted(
                differences,
                key=lambda x: x['abs_z_score'],
                reverse=True
            )
            
            for i, diff in enumerate(ranked_differences):
                diff['rank'] = i + 1
            
            ranking_stats = {
                'total_ranked':        len(ranked_differences),
                'highest_abs_z_score': ranked_differences[0]['abs_z_score']  if ranked_differences else 0.0,
                'lowest_abs_z_score':  ranked_differences[-1]['abs_z_score'] if ranked_differences else 0.0,
                'ranking_timestamp':   datetime.now().isoformat()
            }
            
            result = {
                'success': True,
                'ranked_differences': ranked_differences,
                'ranking_stats': ranking_stats
            }
            
            logger.info(f"✅ {self.name}: Differences ranked successfully")
            logger.info(f"📊 Highest |z|: {ranking_stats['highest_abs_z_score']:.4f}")
            logger.info(f"📊 Lowest  |z|: {ranking_stats['lowest_abs_z_score']:.4f}")
            
            return result
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Error ranking differences: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def get_top_anomalies(self, ranked_differences: List[Dict[str, Any]], top_n: int = 3) -> List[Dict[str, Any]]:
        """Get top N anomalies from ranked differences."""
        logger.info(f"🔍 {self.name}: Getting top {top_n} anomalies")
        
        top_anomalies = ranked_differences[:top_n]
        
        logger.info(f"✅ {self.name}: Retrieved top {len(top_anomalies)} anomalies")
        return top_anomalies
    
    # ── Sequence comparison (SCPI / slow-burn path) ─────────────────────────

    def compare_sequence_profiles(
        self,
        scpi_array: List[float],
        training_matrix_file: str,
        window_size: int,
        weighted_scpi_array: Optional[List[float]] = None,
    ) -> Dict[str, Any]:
        """
        Compare the SCPI vector against the Training Matrix baseline using a
        sequence-scaled standard error (σ / √L).

        Formula per position k:

            z_seq_k = ( SCPI_k − μ_k ) / ( σ_k / √L )

        The √L factor (where L = window_size) amplifies a sustained daily
        deviation: a 2σ/day anomaly across 5 days yields z_seq ≈ 4.47 (HIGH).
        This is the key property that makes slow-burn fraud detectable.

        Deviation bands — identical to single-day path:
            |z_seq| > 3.0  →  HIGH
            |z_seq| > 2.0  →  MEDIUM
            else           →  LOW

        The dollar-weight divergence signal (weighted_scpi_array) mirrors the
        single-day logic: a large |wSCPI − SCPI| means high-value transactions
        are driving the sustained pattern — a compound alarm.

        Returns the same structure as compare_churn_profiles() so downstream
        consumers (anomaly identification, management warnings) need no changes.
        """
        import math
        logger.info(
            f"🔍 {self.name}: Comparing sequence profiles "
            f"(window={window_size})"
        )

        try:
            average_cpi_array = self._load_average_cpi_array(training_matrix_file)
            std_cpi_array     = self._load_std_cpi_array(training_matrix_file)

            if not average_cpi_array:
                return {'success': False, 'error': 'Failed to load μ array from training matrix'}

            if len(scpi_array) != self.cpi_array_size or len(average_cpi_array) != self.cpi_array_size:
                return {
                    'success': False,
                    'error': (
                        f'Array size mismatch: SCPI={len(scpi_array)}, '
                        f'Average={len(average_cpi_array)}, '
                        f'Expected={self.cpi_array_size}'
                    ),
                }

            sqrt_L = math.sqrt(max(window_size, 1))
            differences: List[Dict[str, Any]] = []

            for i in range(len(scpi_array)):
                scpi_val  = scpi_array[i]
                mu_val    = average_cpi_array[i]
                sigma_val = std_cpi_array[i] if i < len(std_cpi_array) else 0.0

                abs_diff = abs(scpi_val - mu_val)

                # Sequence z-score — sigma divided by √L
                if sigma_val > 0:
                    z_seq = (scpi_val - mu_val) / (sigma_val / sqrt_L)
                else:
                    z_seq = 0.0

                abs_z = abs(z_seq)
                if abs_z > 3.0:
                    deviation_level = 'HIGH'
                elif abs_z > 2.0:
                    deviation_level = 'MEDIUM'
                else:
                    deviation_level = 'LOW'

                if mu_val != 0:
                    pct_dev = (abs_diff / abs(mu_val)) * 100
                else:
                    pct_dev = 100.0 if scpi_val != 0 else 0.0

                differences.append({
                    'index':               i + 1,
                    'today_value':         scpi_val,    # SCPI fills the "today" slot
                    'average_value':       mu_val,
                    'std_value':           sigma_val,
                    'absolute_difference': abs_diff,
                    'z_score':             z_seq,
                    'abs_z_score':         abs_z,
                    'percent_deviation':   pct_dev,
                    'deviation_level':     deviation_level,
                    'is_sequence':         True,
                    'window_size':         window_size,
                })

            distance_metrics = self._calculate_array_distance_metrics(scpi_array, average_cpi_array)
            summary_stats    = self._calculate_summary_statistics(differences, distance_metrics)

            scpi_dict   = {str(i + 1): v for i, v in enumerate(scpi_array)}
            avg_dict    = {str(i + 1): v for i, v in enumerate(average_cpi_array)}
            std_dict    = {str(i + 1): v for i, v in enumerate(std_cpi_array)}

            # Dollar-weight signal on the SCPI path — same primary/fallback logic
            has_dollar_weighting = (
                weighted_scpi_array is not None
                and len(weighted_scpi_array) == len(scpi_array)
            )
            dollar_baseline = self._load_dollar_weighted_baseline(training_matrix_file)

            if has_dollar_weighting:
                if dollar_baseline['available']:
                    dollar_signals = self._calculate_dollar_z_score_signal(
                        weighted_scpi_array,
                        dollar_baseline['mu_w'],
                        dollar_baseline['sigma_w'],
                    )
                    dollar_method = 'z_score'
                else:
                    dollar_signals = self._calculate_dollar_weight_signal(
                        scpi_array, weighted_scpi_array
                    )
                    dollar_method = 'divergence_fallback'

                dollar_alarm_count = 0
                for idx, diff in enumerate(differences):
                    sig = dollar_signals[idx] if idx < len(dollar_signals) else {}
                    diff['weighted_cpi']             = weighted_scpi_array[idx]
                    diff['dollar_alarm_level']        = sig.get('dollar_level', 'LOW')
                    diff['dollar_alarm']              = sig.get('dollar_alarm', False)
                    diff['dollar_z_score']            = sig.get('dollar_z_score', 0.0)
                    diff['dollar_abs_z_score']        = sig.get('dollar_abs_z_score', 0.0)
                    diff['dollar_weight_divergence']  = sig.get('divergence', 0.0)
                    diff['dollar_signal_method']      = dollar_method
                    diff['composite_alarm'] = (
                        diff['deviation_level'] in ('HIGH', 'MEDIUM')
                        or diff['dollar_alarm']
                    )
                    if diff['dollar_alarm']:
                        dollar_alarm_count += 1
                logger.info(
                    f"💵 Sequence dollar-weight signal ({dollar_method}): "
                    f"{dollar_alarm_count} position(s) elevated"
                )
            else:
                for diff in differences:
                    diff['composite_alarm'] = diff['deviation_level'] in ('HIGH', 'MEDIUM')

            result: Dict[str, Any] = {
                'success':                   True,
                'is_sequence':               True,
                'window_size':               window_size,
                'today_cpi_array':           scpi_dict,
                'average_cpi_array':         avg_dict,
                'std_cpi_array':             std_dict,
                'differences':               differences,
                'distance_metrics':          distance_metrics,
                'summary_stats':             summary_stats,
                'has_dollar_weighting':      has_dollar_weighting,
                'dollar_baseline_available': dollar_baseline['available'],
                'dollar_baseline_days':      dollar_baseline.get('training_days_w', 0),
                'comparison_timestamp':      datetime.now().isoformat(),
            }
            if has_dollar_weighting:
                result['weighted_cpi_array'] = {
                    str(i + 1): v for i, v in enumerate(weighted_scpi_array)
                }

            high   = summary_stats['high_deviation_count']
            medium = summary_stats['medium_deviation_count']
            logger.info(
                f"✅ {self.name}: Sequence comparison done  "
                f"HIGH={high}  MEDIUM={medium}  "
                f"mean|z_seq|={summary_stats['mean_abs_z_score']:.4f}"
            )
            return result

        except Exception as e:
            logger.error(f"❌ {self.name}: compare_sequence_profiles failed — {e}")
            return {'success': False, 'error': str(e)}

    # ────────────────────────────────────────────────────────────────────────

    def save_comparison_results(self, comparison_result: Dict[str, Any], output_file: str = "churn_comparison_results.json"):
        """Save churn comparison results to file."""
        try:
            with open(output_file, 'w') as f:
                json.dump(comparison_result, f, indent=2, default=str)
            
            logger.info(f"💾 Churn comparison results saved to {output_file}")
            
        except Exception as e:
            logger.error(f"❌ Failed to save churn comparison results: {str(e)}")

def main():
    """Main function for testing the churn comparison agent."""
    print("🚀 NGAME Churn Comparison Agent Test")
    print("=" * 50)
    
    # Initialize agent
    agent = NGameChurnComparisonAgent()
    
    # Create test data
    today_cpi_array = [0.627, 0.521, 0.445, 0.389, 0.334, 0.298, 0.267, 0.234, 0.201, 0.178, 0.156, 0.134, 0.112, 0.090, 0.067, 0.045, 0.023]
    training_matrix_file = "NGAME_Training_Matrix.xlsx"
    
    # Compare churn profiles
    result = agent.compare_churn_profiles(today_cpi_array, training_matrix_file)
    
    if result['success']:
        print(f"\n✅ Churn comparison completed successfully!")
        print(f"📊 Mean |z-score|: {result['summary_stats']['mean_abs_z_score']:.4f}")
        print(f"📊 Max  |z-score|: {result['summary_stats']['max_abs_z_score']:.4f}")
        print(f"📊 High deviations  (|z|>3): {result['summary_stats']['high_deviation_count']}")
        print(f"📊 Medium deviations(|z|>2): {result['summary_stats']['medium_deviation_count']}")
        print(f"📊 Low deviations   (|z|≤2): {result['summary_stats']['low_deviation_count']}")
        
        # Display array distance metrics
        print(f"\n📏 Array Distance Metrics:")
        print(f"   Euclidean distance: {result['distance_metrics']['euclidean_distance']:.6f}")
        print(f"   Manhattan distance: {result['distance_metrics']['manhattan_distance']:.6f}")
        print(f"   Cosine distance: {result['distance_metrics']['cosine_distance']:.6f}")
        print(f"   Cosine similarity: {result['distance_metrics']['cosine_similarity']:.6f}")
        print(f"   Mean Squared Error (MSE): {result['distance_metrics']['mean_squared_error']:.6f}")
        print(f"   Root Mean Squared Error (RMSE): {result['distance_metrics']['root_mean_squared_error']:.6f}")
        
        # Rank differences
        ranking_result = agent.rank_differences(result['differences'])
        
        if ranking_result['success']:
            print(f"\n📊 Top 3 anomalies (by |z-score|):")
            top_anomalies = agent.get_top_anomalies(ranking_result['ranked_differences'], 3)
            for i, anomaly in enumerate(top_anomalies, 1):
                print(f"  {i}. Index {anomaly['index']}: z={anomaly['z_score']:.3f}  ({anomaly['deviation_level']})")
        
        # Save results
        agent.save_comparison_results(result)
        
        return True
    else:
        print(f"\n❌ Churn comparison failed: {result.get('error', 'Unknown error')}")
        return False

if __name__ == "__main__":
    main()