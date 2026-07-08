#!/usr/bin/env python3
"""
NGAME Fraud Analysis Flow Manager
Implements Chart 2: Daily Fraud Analysis Process
Manages the complete daily fraud detection workflow with 7 requirements
"""

import os
import json
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class NGameFraudAnalysisFlowManager:
    """
    NGAME Fraud Analysis Flow Manager for Chart 2 implementation.
    Manages the daily fraud analysis process with 7 requirements.
    """
    
    def __init__(self, sequence_window: int = 5):
        self.training_matrix_file = "NGAME_Training_Matrix.xlsx"
        self.fraud_analysis_file = "NGAME_Fraud_Analysis.json"
        self.asset_misappropriation_taxonomy = "Asset_Misappropriation.ttl"
        self.cpi_array_size = 18  # 18 transaction types (incl. Vendors φ18)

        # Number of days in the rolling SCPI window (configurable)
        self.sequence_window = sequence_window

        # Initialize agents
        self.data_extraction_agent = None
        self.cpi_analysis_agent = None
        self.churn_comparison_agent = None
        self.anomaly_identification_agent = None
        self.account_mapping_agent = None
        self.llm_analysis_agent = None
        self.management_warning_agent = None
        self.sequence_cpi_agent = None
        self.credit_card_watch_agent = None

    def initialize_agents(self):
        """Initialize all agents for fraud analysis flow."""
        logger.info("🤖 Initializing Fraud Analysis Flow Agents...")
        
        # Import and initialize agents
        from ngame_data_extraction_agent import NGameDataExtractionAgent
        from ngame_cpi_analysis_agent import NGameCpiAnalysisAgent
        from ngame_churn_comparison_agent import NGameChurnComparisonAgent
        from ngame_anomaly_identification_agent import NGameAnomalyIdentificationAgent
        from ngame_account_mapping_agent import NGameAccountMappingAgent
        from ngame_llm_analysis_agent import NGameLLMAnalysisAgent
        from ngame_management_warning_agent import NGameManagementWarningAgent
        
        self.data_extraction_agent = NGameDataExtractionAgent()
        self.cpi_analysis_agent = NGameCpiAnalysisAgent()
        self.churn_comparison_agent = NGameChurnComparisonAgent()
        self.anomaly_identification_agent = NGameAnomalyIdentificationAgent()
        self.account_mapping_agent = NGameAccountMappingAgent()
        self.llm_analysis_agent = NGameLLMAnalysisAgent()
        self.management_warning_agent = NGameManagementWarningAgent()

        from ngame_sequence_cpi_agent import NGameSequenceCpiAgent
        self.sequence_cpi_agent = NGameSequenceCpiAgent()

        from ngame_credit_card_watch_agent import NGameCreditCardWatchAgent
        self.credit_card_watch_agent = NGameCreditCardWatchAgent()

        logger.info("✅ All Fraud Analysis Flow agents initialized successfully")
        
    def execute_daily_fraud_analysis(self) -> Dict[str, Any]:
        """
        Execute daily fraud analysis process.
        Implements the 7 requirements from Chart 2.
        """
        logger.info("🔍 Executing Daily Fraud Analysis Process")
        print("🔍 NGAME Daily Fraud Analysis")
        print("=" * 60)
        print("🎯 Analyzing Today's churn for fraud detection")
        
        try:
            # Requirement 1: Execute at close of QuickBooks workday
            print(f"\n📊 Step 1: End-of-Day Execution")
            print(f"⏰ Executing at close of QuickBooks workday: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Requirement 2: Generate new 17-element CPI array (Today's churn profile)
            print(f"\n📊 Step 2: Generate Today's Churn Profile")
            extraction_result = self.data_extraction_agent.extract_daily_data()
            
            if not extraction_result['success']:
                raise Exception(f"Data extraction failed: {extraction_result['error']}")

            # ── Anomaly injection (demo mode) ─────────────────────────────────
            # If run_anomaly_injection_test.py was run beforehand it leaves a flag
            # file.  We inject synthetic vendor URIs into Today.ttl here, *after*
            # extraction but *before* CPI analysis, so the pipeline sees them.
            _injection_flag = "anomaly_injection_active.json"
            if os.path.exists(_injection_flag):
                print(f"\n💉 Anomaly injection flag detected — appending synthetic vendors to Today.ttl")
                try:
                    from ngame_anomaly_injection_agent import NGameAnomalyInjectionAgent
                    _inj_agent = NGameAnomalyInjectionAgent()
                    _ttl_file  = extraction_result.get('ttl_file', 'quickbooks_ontology_Today.ttl')
                    _inj_result = _inj_agent.inject_into_ttl(_ttl_file)
                    if _inj_result.get('success'):
                        print(f"✅ Injected {_inj_result['vendors_injected']} synthetic vendors into {_ttl_file}")
                    else:
                        print(f"⚠️  Injection failed: {_inj_result.get('error')} — proceeding without injection")
                except Exception as _inj_err:
                    print(f"⚠️  Injection error: {_inj_err} — proceeding without injection")
                finally:
                    try:
                        os.remove(_injection_flag)
                    except OSError:
                        pass
            # ─────────────────────────────────────────────────────────────────

            # ── Step 2c: Credit Card Watch (independent; no Training Matrix needed) ──
            print(f"\n💳 Step 2c: Credit Card Watch")
            cc_watch_result: Dict[str, Any] = {}
            if self.credit_card_watch_agent is not None:
                cc_watch_result = self.credit_card_watch_agent.scan_today_transactions()
                cc_sum = cc_watch_result.get('summary', {})
                print(
                    f"✅ CC scan complete — "
                    f"{cc_sum.get('total_cc_transactions', 0)} CC txn(s), "
                    f"{cc_sum.get('total_flagged', 0)} flagged "
                    f"[CRITICAL={cc_sum.get('critical_count', 0)}  "
                    f"HIGH={cc_sum.get('high_count', 0)}  "
                    f"MEDIUM={cc_sum.get('medium_count', 0)}]"
                )
                crit_hi = cc_sum.get('critical_count', 0) + cc_sum.get('high_count', 0)
                if crit_hi > 0:
                    print(
                        f"🚨 CC WATCH ALERT — {crit_hi} high-risk CC transaction(s) detected"
                    )
            else:
                print("   ⚠️  CC watch agent not initialised — skipping")

            cpi_result = self.cpi_analysis_agent.analyze_cpi_coefficients()

            if not cpi_result['success']:
                raise Exception(f"CPI analysis failed: {cpi_result['error']}")

            today_cpi_array          = cpi_result['cpi_array']
            today_weighted_cpi_array = cpi_result.get('weighted_cpi_array')
            print(f"✅ Generated Today's churn profile: {len(today_cpi_array)} transaction types")
            if today_weighted_cpi_array:
                print(f"   Dollar-weighted CPI array computed in parallel")

            # ── Rolling buffer append (feeds the sequence / slow-burn path) ──
            self.cpi_analysis_agent.append_to_rolling_buffer(cpi_result)

            # ── Sequence path (parallel to single-day; silent when buffer < L) ──
            sequence_result: Optional[Dict[str, Any]] = None
            if self.sequence_cpi_agent is not None:
                print(f"\n📊 Step 2b: Sequence CPI Analysis (window={self.sequence_window})")
                scpi_data = self.sequence_cpi_agent.compute_scpi(window=self.sequence_window)
                if scpi_data.get('success') and scpi_data.get('ready'):
                    seq_comparison = self.churn_comparison_agent.compare_sequence_profiles(
                        scpi_data['scpi_array'],
                        self.training_matrix_file,
                        scpi_data['window_size'],
                        weighted_scpi_array=scpi_data.get('weighted_scpi_array'),
                    )
                    sequence_result = {
                        'scpi':       scpi_data,
                        'comparison': seq_comparison,
                    }
                    high   = scpi_data.get('high_count', 0)
                    medium = scpi_data.get('medium_count', 0)
                    print(
                        f"✅ Sequence SCPI computed over {scpi_data['window_size']} days  "
                        f"[HIGH={high}  MEDIUM={medium}]"
                    )
                    if high > 0:
                        print(
                            f"🚨 SLOW-BURN HIGH ALERT — sustained anomaly detected "
                            f"over {scpi_data['window_size']} days"
                        )
                elif scpi_data.get('success'):
                    days_avail = scpi_data.get('days_available', 0)
                    print(
                        f"   ℹ️  Sequence path building up: "
                        f"{days_avail}/{self.sequence_window} days in buffer"
                    )
                else:
                    print(f"   ⚠️  Sequence CPI skipped: {scpi_data.get('error', 'unknown error')}")

            # Requirement 3: Compare Today array with Average CPI array from Training Matrix
            print(f"\n📊 Step 3: Churn Profile Comparison")
            comparison_result = self.churn_comparison_agent.compare_churn_profiles(
                today_cpi_array,
                self.training_matrix_file,
                weighted_cpi_array=today_weighted_cpi_array,
            )
            
            if not comparison_result['success']:
                raise Exception(f"Churn comparison failed: {comparison_result['error']}")
            
            print(f"✅ Compared Today's array with stored Average CPI array")
            
            # Requirement 4: Rank differences for each element pair
            print(f"\n📊 Step 4: Rank Differences")
            ranking_result = self.churn_comparison_agent.rank_differences(
                comparison_result['differences']
            )
            
            if not ranking_result['success']:
                raise Exception(f"Difference ranking failed: {ranking_result['error']}")
            
            print(f"✅ Ranked differences for all 17 element pairs")
            
            # Requirement 5: Find accounts for top 3 most different element pairs
            print(f"\n📊 Step 5: Top 3 Anomaly Identification")
            anomaly_result = self.anomaly_identification_agent.identify_top_anomalies(
                ranking_result['ranked_differences'], top_n=3
            )
            
            if not anomaly_result['success']:
                raise Exception(f"Anomaly identification failed: {anomaly_result['error']}")
            
            print(f"✅ Identified top 3 most different transaction types")
            
            # Map to Chart of Accounts
            account_result = self.account_mapping_agent.map_to_chart_of_accounts(
                anomaly_result['top_anomalies']
            )
            
            if not account_result['success']:
                raise Exception(f"Account mapping failed: {account_result['error']}")
            
            print(f"✅ Mapped to Chart of Accounts for each transaction type")
            
            # Requirement 6: Use RAG/Ollama LLM to match accounts with Misappropriated Assets taxonomy
            print(f"\n📊 Step 6: RAG/LLM Misappropriation Analysis")
            llm_result = self.llm_analysis_agent.analyze_misappropriation_risks(
                account_result['account_mappings'], self.asset_misappropriation_taxonomy
            )
            
            if not llm_result['success']:
                raise Exception(f"LLM analysis failed: {llm_result['error']}")
            
            print(f"✅ LLM analysis completed with {len(llm_result['matches'])} potential matches")
            
            # Requirement 7: Generate management-level warnings based on LLM matches
            print(f"\n📊 Step 7: Management Warning Generation")
            warning_result = self.management_warning_agent.generate_management_warnings(
                llm_result['matches'], llm_result['risk_assessments']
            )
            
            if not warning_result['success']:
                raise Exception(f"Warning generation failed: {warning_result['error']}")
            
            print(f"✅ Generated {len(warning_result['warnings'])} management warnings")
            
            # Compile complete fraud analysis result
            seq_high   = sequence_result['scpi'].get('high_count',   0) if sequence_result else 0
            seq_medium = sequence_result['scpi'].get('medium_count', 0) if sequence_result else 0
            _cc_sum    = cc_watch_result.get('summary', {})
            fraud_analysis_result = {
                'success': True,
                'execution_time':            datetime.now().isoformat(),
                'today_cpi_array':           today_cpi_array,
                'today_weighted_cpi_array':  today_weighted_cpi_array,
                'comparison_result':  comparison_result,
                'ranking_result':     ranking_result,
                'anomaly_result':     anomaly_result,
                'account_result':     account_result,
                'llm_result':         llm_result,
                'warning_result':     warning_result,
                'sequence_result':    sequence_result,   # None when buffer < window
                'cc_watch_result':    cc_watch_result,
                'summary': {
                    'total_transaction_types':       len(today_cpi_array),
                    'top_anomalies_identified':      len(anomaly_result['top_anomalies']),
                    'accounts_analyzed':             len(account_result['account_mappings']),
                    'llm_matches_found':             len(llm_result['matches']),
                    'management_warnings_generated': len(warning_result['warnings']),
                    'overall_risk_level':            warning_result.get('overall_risk_level', 'UNKNOWN'),
                    'sequence_window':               self.sequence_window,
                    'sequence_active':               sequence_result is not None,
                    'sequence_high_count':           seq_high,
                    'sequence_medium_count':         seq_medium,
                    'cc_watch_flagged':              _cc_sum.get('total_flagged', 0),
                    'cc_watch_transactions':         _cc_sum.get('total_cc_transactions', 0),
                    'cc_watch_critical':             _cc_sum.get('critical_count', 0),
                    'cc_watch_high':                 _cc_sum.get('high_count', 0),
                    'cc_watch_structuring':          _cc_sum.get('structuring_flags', 0),
                    'cc_watch_highest_risk':         _cc_sum.get('highest_risk_level', 'CLEAR'),
                }
            }
            
            # Save fraud analysis results
            self.save_fraud_analysis_results(fraud_analysis_result)
            
            logger.info("✅ Daily fraud analysis completed successfully")
            return fraud_analysis_result
            
        except Exception as e:
            logger.error(f"❌ Daily fraud analysis failed: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'execution_time': datetime.now().isoformat()
            }
    
    def save_fraud_analysis_results(self, fraud_analysis_result: Dict[str, Any]):
        """Save fraud analysis results to file."""
        try:
            with open(self.fraud_analysis_file, 'w') as f:
                json.dump(fraud_analysis_result, f, indent=2, default=str)
            
            logger.info(f"💾 Fraud analysis results saved to {self.fraud_analysis_file}")
            
            # Also save a readable version with 1-based indexing
            self.save_readable_fraud_analysis_results(fraud_analysis_result)
            
        except Exception as e:
            logger.error(f"❌ Failed to save fraud analysis results: {str(e)}")
    
    def save_readable_fraud_analysis_results(self, fraud_analysis_result: Dict[str, Any]):
        """Save fraud analysis results with 1-based indexing for readability."""
        try:
            # Create a copy for readable formatting
            readable_result = fraud_analysis_result.copy()
            
            # Convert CPI arrays to 1-based indexing format
            if 'today_cpi_array' in readable_result:
                readable_result['today_cpi_array_readable'] = {}
                for i, value in enumerate(readable_result['today_cpi_array']):
                    cpi_index = i + 1  # Convert to 1-based indexing
                    key = f"φ{cpi_index}"
                    readable_result['today_cpi_array_readable'][key] = value
            
            # Convert comparison results to 1-based indexing
            if 'comparison_result' in readable_result and readable_result['comparison_result']['success']:
                comparison = readable_result['comparison_result']
                if 'today_cpi_array' in comparison:
                    comparison['today_cpi_array_readable'] = {}
                    for i, value in enumerate(comparison['today_cpi_array']):
                        cpi_index = i + 1  # Convert to 1-based indexing
                        key = f"φ{cpi_index}"
                        comparison['today_cpi_array_readable'][key] = value
                if 'average_cpi_array' in comparison:
                    comparison['average_cpi_array_readable'] = {}
                    for i, value in enumerate(comparison['average_cpi_array']):
                        cpi_index = i + 1  # Convert to 1-based indexing
                        key = f"φ{cpi_index}"
                        comparison['average_cpi_array_readable'][key] = value
                
                # Convert differences to 1-based indexing
                if 'differences' in comparison:
                    for diff in comparison['differences']:
                        diff['cpi_index_readable'] = f"φ{diff['index']}"
            
            # Convert top anomalies to 1-based indexing
            if 'anomaly_result' in readable_result and readable_result['anomaly_result']['success']:
                anomaly_result = readable_result['anomaly_result']
                if 'top_anomalies' in anomaly_result:
                    for anomaly in anomaly_result['top_anomalies']:
                        anomaly['cpi_index_readable'] = f"φ{anomaly['index']}"
            
            # Save readable version
            readable_file = self.fraud_analysis_file.replace('.json', '_readable.json')
            with open(readable_file, 'w', encoding='utf-8') as f:
                json.dump(readable_result, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"💾 Readable fraud analysis results saved to {readable_file}")
            
            # Also save a clean version with only 1-based indexing
            self.save_clean_readable_fraud_analysis_results(readable_result)
            
            # Also save a truly clean version with no 0-based arrays at all
            self.save_truly_clean_readable_fraud_analysis_results(readable_result)
            
        except Exception as e:
            logger.error(f"❌ Failed to save readable fraud analysis results: {str(e)}")
    
    def save_clean_readable_fraud_analysis_results(self, readable_result: Dict[str, Any]):
        """Save a clean readable version with only 1-based indexing (no 0-based indices)."""
        try:
            # Create a clean copy
            clean_result = json.loads(json.dumps(readable_result))  # Deep copy
            
            # Clean the differences arrays to use only 1-based indexing
            if 'comparison_result' in clean_result and clean_result['comparison_result']['success']:
                comparison = clean_result['comparison_result']
                
                if 'differences' in comparison:
                    for diff in comparison['differences']:
                        # Remove the 0-based index
                        if 'index' in diff:
                            del diff['index']
                        # Rename cpi_index_readable to just cpi_index for cleaner display
                        if 'cpi_index_readable' in diff:
                            diff['cpi_index'] = diff['cpi_index_readable']
                            del diff['cpi_index_readable']
            
            # Clean the ranking results
            if 'ranking_result' in clean_result and clean_result['ranking_result']['success']:
                ranking = clean_result['ranking_result']
                
                if 'ranked_differences' in ranking:
                    for diff in ranking['ranked_differences']:
                        # Remove the 0-based index
                        if 'index' in diff:
                            del diff['index']
                        # Rename cpi_index_readable to just cpi_index for cleaner display
                        if 'cpi_index_readable' in diff:
                            diff['cpi_index'] = diff['cpi_index_readable']
                            del diff['cpi_index_readable']
            
            # Clean the anomaly results
            if 'anomaly_result' in clean_result and clean_result['anomaly_result']['success']:
                anomaly_result = clean_result['anomaly_result']
                
                if 'top_anomalies' in anomaly_result:
                    for anomaly in anomaly_result['top_anomalies']:
                        # Remove the 0-based index
                        if 'index' in anomaly:
                            del anomaly['index']
                        # Rename cpi_index_readable to just cpi_index for cleaner display
                        if 'cpi_index_readable' in anomaly:
                            anomaly['cpi_index'] = anomaly['cpi_index_readable']
                            del anomaly['cpi_index_readable']
            
            # Clean the account mapping results
            if 'account_result' in clean_result and clean_result['account_result']['success']:
                account_result = clean_result['account_result']
                
                if 'account_mappings' in account_result:
                    for mapping in account_result['account_mappings']:
                        if 'anomaly' in mapping:
                            anomaly = mapping['anomaly']
                            # Remove the 0-based index
                            if 'index' in anomaly:
                                del anomaly['index']
                            # Rename cpi_index_readable to just cpi_index for cleaner display
                            if 'cpi_index_readable' in anomaly:
                                anomaly['cpi_index'] = anomaly['cpi_index_readable']
                                del anomaly['cpi_index_readable']
            
            # Save clean version
            clean_file = self.fraud_analysis_file.replace('.json', '_readable_clean.json')
            with open(clean_file, 'w', encoding='utf-8') as f:
                json.dump(clean_result, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"💾 Clean readable fraud analysis results saved to {clean_file}")
            
        except Exception as e:
            logger.error(f"❌ Failed to save clean readable fraud analysis results: {str(e)}")
    
    def save_truly_clean_readable_fraud_analysis_results(self, readable_result: Dict[str, Any]):
        """Save a truly clean readable version with ONLY 1-based indexing - no 0-based arrays at all."""
        try:
            # Create a truly clean version
            clean_data = {}
            
            # Copy basic metadata
            clean_data['success'] = readable_result['success']
            clean_data['execution_time'] = readable_result['execution_time']
            
            # Replace 0-based arrays with 1-based readable versions
            if 'today_cpi_array_readable' in readable_result:
                clean_data['today_cpi_array'] = readable_result['today_cpi_array_readable']
            
            # Clean the comparison results
            if 'comparison_result' in readable_result and readable_result['comparison_result']['success']:
                comparison = readable_result['comparison_result']
                clean_comparison = {
                    'success': comparison['success'],
                    'comparison_timestamp': comparison.get('comparison_timestamp', '')
                }
                
                # Replace 0-based arrays with 1-based readable versions
                if 'today_cpi_array_readable' in comparison:
                    clean_comparison['today_cpi_array'] = comparison['today_cpi_array_readable']
                if 'average_cpi_array_readable' in comparison:
                    clean_comparison['average_cpi_array'] = comparison['average_cpi_array_readable']
                
                # Clean differences array
                if 'differences' in comparison:
                    clean_differences = []
                    for diff in comparison['differences']:
                        clean_diff = {
                            'cpi_index': f"φ{diff['index']}",
                            'today_value': diff['today_value'],
                            'average_value': diff['average_value'],
                            'absolute_difference': diff['absolute_difference'],
                            'percent_deviation': diff['percent_deviation'],
                            'deviation_level': diff['deviation_level'],
                            'rank': diff['rank']
                        }
                        clean_differences.append(clean_diff)
                    clean_comparison['differences'] = clean_differences
                
                # Copy summary stats
                if 'summary_stats' in comparison:
                    clean_comparison['summary_stats'] = comparison['summary_stats']
                
                clean_data['comparison_result'] = clean_comparison
            
            # Clean the ranking results
            if 'ranking_result' in readable_result and readable_result['ranking_result']['success']:
                ranking = readable_result['ranking_result']
                clean_ranking = {
                    'success': ranking['success'],
                    'ranking_timestamp': ranking.get('ranking_timestamp', '')
                }
                
                # Clean ranked differences array
                if 'ranked_differences' in ranking:
                    clean_ranked_differences = []
                    for diff in ranking['ranked_differences']:
                        clean_diff = {
                            'cpi_index': f"φ{diff['index']}",
                            'today_value': diff['today_value'],
                            'average_value': diff['average_value'],
                            'absolute_difference': diff['absolute_difference'],
                            'percent_deviation': diff['percent_deviation'],
                            'deviation_level': diff['deviation_level'],
                            'rank': diff['rank']
                        }
                        clean_ranked_differences.append(clean_diff)
                    clean_ranking['ranked_differences'] = clean_ranked_differences
                
                # Copy ranking stats
                if 'ranking_stats' in ranking:
                    clean_ranking['ranking_stats'] = ranking['ranking_stats']
                
                clean_data['ranking_result'] = clean_ranking
            
            # Clean the anomaly results
            if 'anomaly_result' in readable_result and readable_result['anomaly_result']['success']:
                anomaly_result = readable_result['anomaly_result']
                clean_anomaly_result = {
                    'success': anomaly_result['success'],
                    'identification_timestamp': anomaly_result.get('identification_timestamp', '')
                }
                
                # Clean top anomalies array
                if 'top_anomalies' in anomaly_result:
                    clean_top_anomalies = []
                    for anomaly in anomaly_result['top_anomalies']:
                        clean_anomaly = {
                            'cpi_index': f"φ{anomaly['index']}",
                            'transaction_type': anomaly['transaction_type'],
                            'today_value': anomaly['today_value'],
                            'average_value': anomaly['average_value'],
                            'absolute_difference': anomaly['absolute_difference'],
                            'percent_deviation': anomaly['percent_deviation'],
                            'deviation_level': anomaly['deviation_level'],
                            'rank': anomaly['rank']
                        }
                        clean_top_anomalies.append(clean_anomaly)
                    clean_anomaly_result['top_anomalies'] = clean_top_anomalies
                
                # Copy anomaly stats
                if 'anomaly_stats' in anomaly_result:
                    clean_anomaly_result['anomaly_stats'] = anomaly_result['anomaly_stats']
                
                clean_data['anomaly_result'] = clean_anomaly_result
            
            # Copy other results as-is (they don't contain 0-based arrays)
            for key in ['account_result', 'llm_result', 'warning_result', 'summary']:
                if key in readable_result:
                    clean_data[key] = readable_result[key]

            # Credit Card Watch — keep summary + top flags for dashboard display
            if readable_result.get('cc_watch_result'):
                ccwr = readable_result['cc_watch_result']
                clean_data['cc_watch_result'] = {
                    'success': ccwr.get('success'),
                    'summary': ccwr.get('summary') or {},
                    'cc_flags': (ccwr.get('cc_flags') or [])[:20],
                    'scan_timestamp': ccwr.get('scan_timestamp'),
                }
            
            # Save truly clean version
            clean_file = self.fraud_analysis_file.replace('.json', '_readable_truly_clean.json')
            with open(clean_file, 'w', encoding='utf-8') as f:
                json.dump(clean_data, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"💾 Truly clean readable fraud analysis results saved to {clean_file}")
            
        except Exception as e:
            logger.error(f"❌ Failed to save truly clean readable fraud analysis results: {str(e)}")
    
    def show_fraud_analysis_summary(self, fraud_analysis_result: Dict[str, Any]):
        """Show fraud analysis summary."""
        if not fraud_analysis_result['success']:
            print(f"❌ Fraud analysis failed: {fraud_analysis_result.get('error', 'Unknown error')}")
            return
        
        print(f"\n📊 Daily Fraud Analysis Summary")
        print("=" * 50)
        
        summary = fraud_analysis_result['summary']
        
        print(f"⏰ Execution Time: {fraud_analysis_result['execution_time']}")
        print(f"📊 Transaction Types Analyzed: {summary['total_transaction_types']}")
        print(f"🔍 Top Anomalies Identified: {summary['top_anomalies_identified']}")
        print(f"📋 Accounts Analyzed: {summary['accounts_analyzed']}")
        print(f"🤖 LLM Matches Found: {summary['llm_matches_found']}")
        print(f"⚠️  Management Warnings: {summary['management_warnings_generated']}")
        print(f"🎯 Overall Risk Level: {summary['overall_risk_level']}")

        # Sequence path summary
        if summary.get('sequence_active'):
            print(f"\n🔁 Sequence SCPI (window={summary['sequence_window']})")
            print(f"   HIGH positions:   {summary['sequence_high_count']}")
            print(f"   MEDIUM positions: {summary['sequence_medium_count']}")
            if summary['sequence_high_count'] > 0:
                print("   🚨 SLOW-BURN ALERT: sustained anomaly detected across window")
        else:
            days_in_buffer = 0
            if fraud_analysis_result.get('sequence_result') is None:
                from ngame_sequence_cpi_agent import NGameSequenceCpiAgent
                days_in_buffer = len(NGameSequenceCpiAgent().get_buffer_summary().get('days_stored', 0)
                                     if False else [])
            print(
                f"\n🔁 Sequence SCPI: building buffer "
                f"(window={summary['sequence_window']})"
            )
        
        # CC Watch summary
        if summary.get('cc_watch_flagged', 0) > 0 or summary.get('cc_watch_critical', 0) > 0:
            print(f"\n💳 Credit Card Watch")
            print(f"   Flagged transactions  : {summary.get('cc_watch_flagged', 0)}")
            print(f"   CRITICAL              : {summary.get('cc_watch_critical', 0)}")
            print(f"   HIGH                  : {summary.get('cc_watch_high', 0)}")
            print(f"   Structuring patterns  : {summary.get('cc_watch_structuring', 0)}")
            print(f"   Highest risk level    : {summary.get('cc_watch_highest_risk', 'CLEAR')}")
            if summary.get('cc_watch_critical', 0) > 0:
                print("   🚨 CRITICAL — CC misuse with virtually-certain personal use detected")
            elif summary.get('cc_watch_high', 0) > 0:
                print("   ⚠️  HIGH — Strong personal-use signal on company card(s)")
            # Show top flagged vendors if available
            cc_flags = fraud_analysis_result.get('cc_watch_result', {}).get('cc_flags', [])
            if cc_flags:
                print(f"   Top flagged vendors:")
                for flag in cc_flags[:3]:
                    print(
                        f"     [{flag['overall_risk']:8s}] "
                        f"${flag['amount']:>9,.2f}  "
                        f"{flag['vendor_name'][:35]}  GL: {flag['gl_account'][:25]}"
                    )
        else:
            print(f"\n💳 Credit Card Watch: no high-risk CC transactions detected")

        # Show risk interpretation
        self.show_risk_interpretation(summary['overall_risk_level'])
        
        # Show CPI arrays with 1-based indexing for readability
        if 'today_cpi_array' in fraud_analysis_result:
            self.show_cpi_array_readable(fraud_analysis_result['today_cpi_array'], "Today's CPI Array")
        
        # Show CPI comparison if available
        if 'comparison_result' in fraud_analysis_result and fraud_analysis_result['comparison_result']['success']:
            comparison = fraud_analysis_result['comparison_result']
            if 'today_cpi_array' in comparison and 'average_cpi_array' in comparison:
                self.show_cpi_comparison_readable(comparison['today_cpi_array'], comparison['average_cpi_array'])
        
        # Show top anomalies
        if 'anomaly_result' in fraud_analysis_result:
            self.show_top_anomalies(fraud_analysis_result['anomaly_result'])
        
        # Show management warnings
        if 'warning_result' in fraud_analysis_result:
            self.show_management_warnings(fraud_analysis_result['warning_result'])
    
    def show_risk_interpretation(self, risk_level: str):
        """Show risk interpretation and recommendations."""
        print(f"\n🎯 Risk Interpretation")
        print("-" * 30)
        
        if risk_level == "HIGH":
            print("🚨 HIGH RISK DETECTED")
            print("   - Significant misappropriation indicators found")
            print("   - Immediate investigation required")
            print("   - Review all flagged accounts and transactions")
        elif risk_level == "MEDIUM":
            print("⚠️  MEDIUM RISK DETECTED")
            print("   - Moderate misappropriation indicators found")
            print("   - Enhanced monitoring recommended")
            print("   - Review specific flagged accounts")
        elif risk_level == "LOW":
            print("✅ LOW RISK DETECTED")
            print("   - Minimal misappropriation indicators found")
            print("   - Normal operation")
            print("   - Continue regular monitoring")
        else:
            print("❓ UNKNOWN RISK LEVEL")
            print("   - Unable to determine risk level")
            print("   - Review analysis results manually")
    
    def show_top_anomalies(self, anomaly_result: Dict[str, Any]):
        """Show top anomalies identified."""
        print(f"\n🔍 Top Anomalies Identified")
        print("-" * 30)
        
        for i, anomaly in enumerate(anomaly_result['top_anomalies'], 1):
            transaction_type = anomaly['transaction_type']
            difference = anomaly['absolute_difference']
            z_score = anomaly.get('z_score', 0.0)
            # 'index' is already 1-based (set by churn comparison agent)
            cpi_index = anomaly['index']
            
            print(f"{i}. {transaction_type} (CPI{cpi_index})")
            print(f"   Difference: {difference:.6f}")
            print(f"   z-score:    {z_score:.3f}")
    
    def show_cpi_array_readable(self, cpi_array, title: str = "CPI Array"):
        """Display CPI array with 1-based indexing for better readability."""
        # compare_churn_profiles returns a {str: float} dict; normalise to list
        if isinstance(cpi_array, dict):
            cpi_array = [cpi_array[k] for k in sorted(cpi_array, key=lambda x: int(x))]

        print(f"\n📊 {title} (1-based indexing)")
        print("-" * 40)

        for i, value in enumerate(cpi_array):
            cpi_index = i + 1
            print(f"CPI{cpi_index:2d} = {value:8.6f}")
    
    def show_cpi_comparison_readable(self, today_cpi, average_cpi):
        """Display CPI array comparison with 1-based indexing."""
        # compare_churn_profiles returns {str: float} dicts; normalise to lists
        if isinstance(today_cpi, dict):
            today_cpi   = [today_cpi[k]   for k in sorted(today_cpi,   key=lambda x: int(x))]
        if isinstance(average_cpi, dict):
            average_cpi = [average_cpi[k] for k in sorted(average_cpi, key=lambda x: int(x))]

        print(f"\n📊 CPI Array Comparison (1-based indexing)")
        print("-" * 60)
        print(f"{'Index':>6} {'Today':>10} {'μ (Avg)':>10} {'Difference':>12} {'% Dev':>8}")
        print("-" * 60)

        for i, (today_val, avg_val) in enumerate(zip(today_cpi, average_cpi)):
            cpi_index = i + 1
            difference = float(today_val) - float(avg_val)
            percent_dev = (difference / float(avg_val) * 100) if avg_val != 0 else 0

            print(f"CPI{cpi_index:2d}  {today_val:10.6f} {avg_val:10.6f} {difference:12.6f} {percent_dev:7.2f}%")

    def show_management_warnings(self, warning_result: Dict[str, Any]):
        """Show management warnings generated."""
        print(f"\n⚠️  Management Warnings")
        print("-" * 30)
        
        warnings = warning_result.get('warnings', [])
        
        if not warnings:
            print("✅ No management warnings generated")
            return
        
        for i, warning in enumerate(warnings, 1):
            account_name = warning.get('account_name', 'Unknown')
            risk_level = warning.get('risk_level', 'Unknown')
            misappropriation_type = warning.get('misappropriation_type', 'Unknown')
            confidence = warning.get('confidence', 0)
            
            print(f"{i}. Account: {account_name}")
            print(f"   Risk Level: {risk_level}")
            print(f"   Misappropriation Type: {misappropriation_type}")
            print(f"   Confidence: {confidence:.2f}")
            print()
    
    def validate_prerequisites(self) -> bool:
        """Validate prerequisites for fraud analysis."""
        print(f"\n🔍 Validating Prerequisites")
        print("-" * 30)
        
        # Check if training matrix exists
        if not os.path.exists(self.training_matrix_file):
            print(f"❌ Training matrix not found: {self.training_matrix_file}")
            print(f"📊 Please complete training period first")
            return False
        
        # Check if asset misappropriation taxonomy exists
        if not os.path.exists(self.asset_misappropriation_taxonomy):
            print(f"❌ Asset misappropriation taxonomy not found: {self.asset_misappropriation_taxonomy}")
            print(f"📊 Please ensure taxonomy file exists")
            return False
        
        # Check training matrix completeness
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(self.training_matrix_file)
            worksheet = workbook.active
            
            # Count training days (A=type, B=μ, C=σ → 3 fixed cols)
            training_days = worksheet.max_column - 3
            
            if training_days < 30:
                print(f"⚠️  Training period incomplete: {training_days}/30 days")
                print(f"📊 Consider completing training period first")
                return False
            
            print(f"✅ Training matrix found with {training_days} days")
            print(f"✅ Asset misappropriation taxonomy found")
            print(f"✅ Prerequisites validated")
            return True
            
        except Exception as e:
            print(f"❌ Error validating training matrix: {e}")
            return False

def main():
    """Main function for NGAME Fraud Analysis Flow Manager."""
    print("🚀 NGAME Fraud Analysis Flow Manager")
    print("=" * 60)
    print("📊 Chart 2: Daily Fraud Analysis Process")
    
    # Initialize fraud analysis flow manager
    fraud_manager = NGameFraudAnalysisFlowManager()
    
    # Validate prerequisites
    if not fraud_manager.validate_prerequisites():
        print(f"\n❌ Prerequisites not met - cannot proceed with fraud analysis")
        return False
    
    # Initialize agents
    fraud_manager.initialize_agents()
    
    # Execute daily fraud analysis
    result = fraud_manager.execute_daily_fraud_analysis()
    
    # Show summary
    fraud_manager.show_fraud_analysis_summary(result)
    
    if result['success']:
        print(f"\n🎉 Daily fraud analysis completed successfully!")
        print(f"📁 Results saved to: {fraud_manager.fraud_analysis_file}")
        return True
    else:
        print(f"\n❌ Daily fraud analysis failed")
        return False

if __name__ == "__main__":
    main()
