#!/usr/bin/env python3
"""
NGAME Matrix Management Agent
Manages the 17×N training matrix for iterative CPI array accumulation
"""

import os
import json
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging

# Handle openpyxl import with fallback
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    print(f"⚠️  Warning: openpyxl not available: {e}")
    print("📦 Installing openpyxl...")
    import subprocess
    import sys
    try:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        OPENPYXL_AVAILABLE = True
        print("✅ openpyxl installed successfully")
    except Exception as install_error:
        print(f"❌ Failed to install openpyxl: {install_error}")
        OPENPYXL_AVAILABLE = False

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class NGameMatrixManagementAgent:
    """
    NGAME Matrix Management Agent for training flow.
    Manages the 17×N training matrix for iterative CPI array accumulation.
    """
    
    def __init__(self):
        self.name = "NGameMatrixManagementAgent"
        self.matrix_file = "NGAME_Training_Matrix.xlsx"

        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            logger.error("❌ openpyxl is not available. Matrix operations will not work.")
            raise ImportError("openpyxl is required for matrix management operations")

        from ngame_transaction_types import STANDARD_TRANSACTION_TYPES, CPI_ARRAY_SIZE

        # Define the 18 standard transaction types in order
        self.transaction_types = list(STANDARD_TRANSACTION_TYPES)
        self.cpi_array_size = CPI_ARRAY_SIZE
        
        logger.info(f"🤖 {self.name} initialized")
    
    def update_training_matrix(self, day_number: int, cpi_array: Optional[List[float]]) -> Dict[str, Any]:
        """
        Update training matrix with new CPI array.
        Implements the matrix evolution from Chart 1.
        """
        logger.info(f"📊 {self.name}: Updating training matrix for Day {day_number}")
        
        try:
            # Check if matrix file exists
            if not os.path.exists(self.matrix_file):
                logger.info(f"📊 {self.name}: Creating initial training matrix")
                return self._create_initial_matrix(day_number, cpi_array)
            
            # Load existing matrix
            workbook = load_workbook(self.matrix_file)
            worksheet = workbook.active
            
            # Check if this is the first day (no CPI array)
            if day_number == 1 and cpi_array is None:
                logger.info(f"ℹ️  {self.name}: Day 1 - no CPI array to add (first run)")
                return {
                    'success': True,
                    'day_number': day_number,
                    'matrix_updated': False,
                    'message': 'Day 1 - no CPI array to add (first run)',
                    'matrix_dimensions': f"17×{day_number}"
                }
            
            # Validate CPI array
            if cpi_array is None:
                return {
                    'success': False,
                    'error': f'No CPI array provided for Day {day_number}'
                }
            
            if len(cpi_array) != self.cpi_array_size:
                return {
                    'success': False,
                    'error': f'Invalid CPI array size: {len(cpi_array)}, expected {self.cpi_array_size}'
                }
            
            # Add new column for this day
            new_column = worksheet.max_column + 1
            
            # Add day header
            worksheet.cell(row=1, column=new_column, value=f"Day {day_number}")
            
            # Add CPI values
            for i, cpi_value in enumerate(cpi_array):
                row = i + 2  # Row 2 is first data row (row 1 is header)
                worksheet.cell(row=row, column=new_column, value=cpi_value)
            
            # Recalculate average CPI values
            self._recalculate_average_cpi(worksheet)
            
            # Apply formatting
            self._apply_matrix_formatting(worksheet)
            
            # Save workbook
            workbook.save(self.matrix_file)
            
            # Get matrix statistics
            matrix_stats = self._get_matrix_statistics(worksheet)
            
            result = {
                'success': True,
                'day_number': day_number,
                'matrix_updated': True,
                'matrix_dimensions': f"17×{day_number}",
                'cpi_array_added': cpi_array,
                'matrix_stats': matrix_stats,
                'update_timestamp': datetime.now().isoformat()
            }
            
            logger.info(f"✅ {self.name}: Training matrix updated successfully")
            logger.info(f"📊 Matrix dimensions: {result['matrix_dimensions']}")
            
            return result
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Error updating training matrix: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _create_initial_matrix(self, day_number: int, cpi_array: Optional[List[float]]) -> Dict[str, Any]:
        """Create initial training matrix with μ (col B), σ (col C), and Day columns from col D."""
        logger.info(f"📊 {self.name}: Creating initial training matrix")
        
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Training Matrix"
            
            # A=Transaction Type, B=μ, C=σ, D=Day 1
            headers = ["Transaction Type", "μ (Mean CPI)", "σ (Std Dev CPI)", "Day 1"]
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for i, tx_type in enumerate(self.transaction_types):
                row = i + 2
                worksheet.cell(row=row, column=1, value=tx_type)
            
            if cpi_array is not None:
                for i, cpi_value in enumerate(cpi_array):
                    row = i + 2
                    worksheet.cell(row=row, column=4, value=cpi_value)  # Day 1 now in col D
                
                self._recalculate_average_cpi(worksheet)
            else:
                for i in range(len(self.transaction_types)):
                    row = i + 2
                    worksheet.cell(row=row, column=2, value=0.0)
                    worksheet.cell(row=row, column=3, value=0.0)
            
            # Apply formatting
            self._apply_matrix_formatting(worksheet)
            
            # Save workbook
            workbook.save(self.matrix_file)
            
            # Get matrix statistics
            matrix_stats = self._get_matrix_statistics(worksheet)
            
            result = {
                'success': True,
                'day_number': day_number,
                'matrix_updated': True,
                'matrix_dimensions': f"17×{day_number}",
                'cpi_array_added': cpi_array,
                'matrix_stats': matrix_stats,
                'update_timestamp': datetime.now().isoformat()
            }
            
            logger.info(f"✅ {self.name}: Initial training matrix created successfully")
            return result
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Error creating initial matrix: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _create_initial_matrix_start_mode(self) -> Dict[str, Any]:
        """
        Create initial training matrix for first-time setup.
        Creates matrix with Transaction Type, μ, and σ columns only (no daily columns yet).
        """
        logger.info(f"📊 {self.name}: Creating first-time training matrix (start mode)")
        
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Training Matrix"
            
            # A=Transaction Type, B=μ, C=σ  (no Day columns yet)
            headers = ["Transaction Type", "μ (Mean CPI)", "σ (Std Dev CPI)"]
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for i, tx_type in enumerate(self.transaction_types):
                row = i + 2
                worksheet.cell(row=row, column=1, value=tx_type)
            
            for i in range(len(self.transaction_types)):
                row = i + 2
                worksheet.cell(row=row, column=2, value=0.0)
                worksheet.cell(row=row, column=3, value=0.0)
            
            # Apply formatting
            self._apply_matrix_formatting(worksheet)
            
            # Save workbook
            workbook.save(self.matrix_file)
            
            # Get matrix statistics
            matrix_stats = self._get_matrix_statistics(worksheet)
            
            result = {
                'success': True,
                'day_number': 0,  # No training days yet
                'matrix_updated': True,
                'matrix_dimensions': "17×0",  # No daily columns yet
                'cpi_array_added': None,
                'matrix_stats': matrix_stats,
                'update_timestamp': datetime.now().isoformat()
            }
            
            logger.info(f"✅ {self.name}: First-time training matrix created successfully")
            logger.info(f"📊 Matrix has only Transaction Type, μ, and σ columns")
            return result
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Error creating first-time matrix: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _recalculate_average_cpi(self, worksheet):
        """
        Recalculate μ (col B) and σ (col C) for all transaction type rows.
        Day columns start at col D (index 4): A=type, B=μ, C=σ, D+=daily CPI.
        """
        logger.info(f"📊 {self.name}: Recalculating μ and σ columns")
        
        # Day columns start at col 4
        data_columns = [
            col for col in range(4, worksheet.max_column + 1)
            if worksheet.cell(row=1, column=col).value
            and str(worksheet.cell(row=1, column=col).value).startswith("Day ")
        ]
        
        if not data_columns:
            logger.warning(f"⚠️  {self.name}: No Day columns found for μ/σ calculation")
            return
        
        for row in range(2, worksheet.max_row + 1):
            cpi_values = []
            for col in data_columns:
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None:
                    try:
                        cpi_values.append(float(cell_value))
                    except (ValueError, TypeError):
                        logger.warning(f"⚠️  Invalid CPI value in row {row}, col {col}: {cell_value}")
            
            if cpi_values:
                mean_cpi = float(np.mean(cpi_values))
                std_cpi  = float(np.std(cpi_values, ddof=1)) if len(cpi_values) >= 2 else 0.0
            else:
                mean_cpi = 0.0
                std_cpi  = 0.0
            
            worksheet.cell(row=row, column=2, value=mean_cpi)
            worksheet.cell(row=row, column=3, value=std_cpi)
        
        logger.info(f"✅ {self.name}: μ and σ columns recalculated")
    
    def _apply_matrix_formatting(self, worksheet):
        """Apply formatting to the training matrix."""
        logger.info(f"🎨 {self.name}: Applying matrix formatting")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(name="Arial", size=10)
        data_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Format data rows
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Auto-adjust column widths
        for col in range(1, worksheet.max_column + 1):
            column_letter = worksheet.cell(row=1, column=col).column_letter
            worksheet.column_dimensions[column_letter].width = 20
        
        logger.info(f"✅ {self.name}: Matrix formatting applied")
    
    def _get_matrix_statistics(self, worksheet) -> Dict[str, Any]:
        """Get statistics for the training matrix."""
        logger.info(f"📊 {self.name}: Calculating matrix statistics")
        
        # Count training days (A=type, B=μ, C=σ → Day columns start at col 4)
        training_days = worksheet.max_column - 3
        
        # Get average CPI values
        average_cpi_values = []
        for row in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row, column=2).value
            if cell_value is not None:
                try:
                    average_cpi_values.append(float(cell_value))
                except (ValueError, TypeError):
                    average_cpi_values.append(0.0)
        
        # Calculate statistics
        if average_cpi_values:
            stats = {
                'training_days': training_days,
                'transaction_types': len(self.transaction_types),
                'matrix_dimensions': f"17×{training_days}",
                'average_cpi_mean': float(np.mean(average_cpi_values)),
                'average_cpi_std': float(np.std(average_cpi_values)),
                'average_cpi_min': float(np.min(average_cpi_values)),
                'average_cpi_max': float(np.max(average_cpi_values)),
                'average_cpi_range': float(np.max(average_cpi_values) - np.min(average_cpi_values))
            }
        else:
            stats = {
                'training_days': training_days,
                'transaction_types': len(self.transaction_types),
                'matrix_dimensions': f"17×{training_days}",
                'average_cpi_mean': 0.0,
                'average_cpi_std': 0.0,
                'average_cpi_min': 0.0,
                'average_cpi_max': 0.0,
                'average_cpi_range': 0.0
            }
        
        logger.info(f"✅ {self.name}: Matrix statistics calculated")
        return stats
    
    def get_matrix_status(self) -> Dict[str, Any]:
        """Get current matrix status."""
        logger.info(f"📊 {self.name}: Getting matrix status")
        
        status = {
            'matrix_file': self.matrix_file,
            'matrix_exists': os.path.exists(self.matrix_file),
            'matrix_dimensions': 'N/A',
            'training_days': 0,
            'training_complete': False
        }
        
        if status['matrix_exists']:
            try:
                workbook = load_workbook(self.matrix_file)
                worksheet = workbook.active
                
                # A=type, B=μ, C=σ → Day columns start at col 4
                training_days = worksheet.max_column - 3
                status['training_days'] = training_days
                status['matrix_dimensions'] = f"17×{training_days}"
                status['training_complete'] = training_days >= 30
                
                status['matrix_stats'] = self._get_matrix_statistics(worksheet)
                
            except Exception as e:
                status['error'] = str(e)
        
        logger.info(f"✅ {self.name}: Matrix status retrieved")
        return status
    
    def get_average_cpi_array(self) -> Optional[List[float]]:
        """Get the current average CPI array from the matrix."""
        logger.info(f"📊 {self.name}: Getting average CPI array")
        
        if not os.path.exists(self.matrix_file):
            logger.warning(f"⚠️  {self.name}: Matrix file not found")
            return None
        
        try:
            workbook = load_workbook(self.matrix_file)
            worksheet = workbook.active
            
            # Extract average CPI values from column 2
            average_cpi_array = []
            for row in range(2, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=row, column=2).value
                if cell_value is not None:
                    try:
                        average_cpi_array.append(float(cell_value))
                    except (ValueError, TypeError):
                        average_cpi_array.append(0.0)
                else:
                    average_cpi_array.append(0.0)
            
            logger.info(f"✅ {self.name}: Retrieved average CPI array with {len(average_cpi_array)} elements")
            return average_cpi_array
            
        except Exception as e:
            logger.error(f"❌ {self.name}: Error getting average CPI array: {e}")
            return None
    
    # ── Dollar-Weighted Baseline (second sheet: "wCPI_Baseline") ────────────
    #
    # Schema mirrors the primary Training Matrix sheet:
    #   Col A  : Transaction Type
    #   Col B  : μ_w  (mean dollar-weighted CPI across training days)
    #   Col C  : σ_w  (sample std dev of dollar-weighted CPI)
    #   Col D+ : wDay 1, wDay 2, …  (daily wCPI values)
    #
    # Stored in a second sheet of the same NGAME_Training_Matrix.xlsx file so
    # the primary sheet schema (and all readers that rely on max_column - 3) is
    # completely unaffected.

    _WCPI_SHEET = "wCPI_Baseline"

    def update_dollar_weighted_baseline(
        self,
        day_number: int,
        weighted_cpi_array: Optional[List[float]],
    ) -> Dict[str, Any]:
        """
        Append today's dollar-weighted CPI to the 'wCPI_Baseline' sheet and
        recompute μ_w / σ_w for every transaction type.

        Called once per training day immediately after update_training_matrix().
        Silently skips when weighted_cpi_array is None or wrong length.
        """
        logger.info(
            f"📊 {self.name}: Updating dollar-weighted baseline for Day {day_number}"
        )

        if weighted_cpi_array is None:
            logger.warning(
                f"⚠️  {self.name}: No weighted_cpi_array for Day {day_number} — "
                f"dollar-weighted baseline not updated"
            )
            return {'success': False, 'error': 'weighted_cpi_array is None'}

        if len(weighted_cpi_array) != self.cpi_array_size:
            return {
                'success': False,
                'error': (
                    f'weighted_cpi_array length {len(weighted_cpi_array)} '
                    f'≠ expected {self.cpi_array_size}'
                ),
            }

        try:
            if os.path.exists(self.matrix_file):
                workbook = load_workbook(self.matrix_file)
            else:
                workbook = Workbook()

            # Create or load the wCPI sheet
            if self._WCPI_SHEET not in workbook.sheetnames:
                ws_w = workbook.create_sheet(self._WCPI_SHEET)
                self._init_wcpi_sheet(ws_w)
            else:
                ws_w = workbook[self._WCPI_SHEET]

            # Append the new day column
            new_col = ws_w.max_column + 1
            ws_w.cell(row=1, column=new_col, value=f"wDay {day_number}")
            ws_w.cell(row=1, column=new_col).font = Font(bold=True, color="FFFFFF")
            ws_w.cell(row=1, column=new_col).fill = PatternFill(
                start_color="1F5C2E", end_color="1F5C2E", fill_type="solid"
            )
            ws_w.cell(row=1, column=new_col).alignment = Alignment(horizontal="center")

            for i, w_val in enumerate(weighted_cpi_array):
                row = i + 2
                ws_w.cell(row=row, column=new_col, value=w_val)

            # Recompute μ_w (col B) and σ_w (col C) across all wDay columns
            self._recalculate_wcpi_stats(ws_w)

            workbook.save(self.matrix_file)

            logger.info(
                f"✅ {self.name}: Dollar-weighted baseline updated "
                f"(Day {day_number})"
            )
            return {
                'success': True,
                'day_number': day_number,
                'sheet': self._WCPI_SHEET,
            }

        except Exception as e:
            logger.error(
                f"❌ {self.name}: Error updating dollar-weighted baseline: {e}"
            )
            return {'success': False, 'error': str(e)}

    def load_dollar_weighted_baseline(self) -> Dict[str, Any]:
        """
        Load μ_w and σ_w arrays from the 'wCPI_Baseline' sheet.

        Returns a dict with keys:
            available         bool  — False when the sheet doesn't exist or
                                      fewer than 2 training days are present
            mu_w              list  — 18-element μ_w array
            sigma_w           list  — 18-element σ_w array
            training_days_w   int   — number of wDay columns present
        """
        zeros = [0.0] * self.cpi_array_size

        if not os.path.exists(self.matrix_file):
            return {'available': False, 'mu_w': zeros, 'sigma_w': zeros,
                    'training_days_w': 0}

        try:
            workbook = load_workbook(self.matrix_file, read_only=True)
            if self._WCPI_SHEET not in workbook.sheetnames:
                return {'available': False, 'mu_w': zeros, 'sigma_w': zeros,
                        'training_days_w': 0}

            ws_w = workbook[self._WCPI_SHEET]

            # Count wDay columns (col D onward)
            wday_cols = [
                c for c in range(4, ws_w.max_column + 1)
                if ws_w.cell(row=1, column=c).value
                and str(ws_w.cell(row=1, column=c).value).startswith("wDay ")
            ]
            training_days_w = len(wday_cols)

            if training_days_w < 2:
                return {
                    'available': False,
                    'mu_w': zeros,
                    'sigma_w': zeros,
                    'training_days_w': training_days_w,
                }

            mu_w    = []
            sigma_w = []
            for row in range(2, ws_w.max_row + 1):
                mu_cell    = ws_w.cell(row=row, column=2).value
                sigma_cell = ws_w.cell(row=row, column=3).value
                try:
                    mu_w.append(float(mu_cell) if mu_cell is not None else 0.0)
                except (ValueError, TypeError):
                    mu_w.append(0.0)
                try:
                    sigma_w.append(float(sigma_cell) if sigma_cell is not None else 0.0)
                except (ValueError, TypeError):
                    sigma_w.append(0.0)

            return {
                'available': True,
                'mu_w': mu_w,
                'sigma_w': sigma_w,
                'training_days_w': training_days_w,
            }

        except Exception as e:
            logger.error(
                f"❌ {self.name}: Error loading dollar-weighted baseline: {e}"
            )
            return {'available': False, 'mu_w': zeros, 'sigma_w': zeros,
                    'training_days_w': 0, 'error': str(e)}

    def _init_wcpi_sheet(self, ws_w):
        """Create the wCPI_Baseline sheet header and transaction-type rows."""
        headers = [
            "Transaction Type",
            "μ_w (Mean wCPI)",
            "σ_w (Std Dev wCPI)",
        ]
        green_fill = PatternFill(
            start_color="1F5C2E", end_color="1F5C2E", fill_type="solid"
        )
        for col, hdr in enumerate(headers, 1):
            cell = ws_w.cell(row=1, column=col, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = green_fill
            cell.alignment = Alignment(horizontal="center")

        for i, tx_type in enumerate(self.transaction_types):
            ws_w.cell(row=i + 2, column=1, value=tx_type)
            ws_w.cell(row=i + 2, column=2, value=0.0)
            ws_w.cell(row=i + 2, column=3, value=0.0)

    def _recalculate_wcpi_stats(self, ws_w):
        """Recompute μ_w (col B) and σ_w (col C) from all wDay columns."""
        wday_cols = [
            c for c in range(4, ws_w.max_column + 1)
            if ws_w.cell(row=1, column=c).value
            and str(ws_w.cell(row=1, column=c).value).startswith("wDay ")
        ]
        if not wday_cols:
            return

        for row in range(2, ws_w.max_row + 1):
            vals = []
            for col in wday_cols:
                v = ws_w.cell(row=row, column=col).value
                if v is not None:
                    try:
                        vals.append(float(v))
                    except (ValueError, TypeError):
                        pass
            mu    = float(np.mean(vals)) if vals else 0.0
            sigma = float(np.std(vals, ddof=1)) if len(vals) >= 2 else 0.0
            ws_w.cell(row=row, column=2, value=mu)
            ws_w.cell(row=row, column=3, value=sigma)

    # ────────────────────────────────────────────────────────────────────────

    def save_matrix_results(self, matrix_result: Dict[str, Any], output_file: str = "matrix_management_results.json"):
        """Save matrix management results to file."""
        try:
            with open(output_file, 'w') as f:
                json.dump(matrix_result, f, indent=2, default=str)
            
            logger.info(f"💾 Matrix management results saved to {output_file}")
            
        except Exception as e:
            logger.error(f"❌ Failed to save matrix management results: {str(e)}")

def main():
    """Main function for testing the matrix management agent."""
    print("🚀 NGAME Matrix Management Agent Test")
    print("=" * 50)
    
    # Initialize agent
    agent = NGameMatrixManagementAgent()
    
    # Test matrix status
    status = agent.get_matrix_status()
    print(f"\n📊 Matrix Status:")
    print(f"   Matrix exists: {status['matrix_exists']}")
    print(f"   Training days: {status['training_days']}")
    print(f"   Matrix dimensions: {status['matrix_dimensions']}")
    print(f"   Training complete: {status['training_complete']}")
    
    # Test with sample CPI array
    test_cpi_array = [0.627, 0.521, 0.445, 0.389, 0.334, 0.298, 0.267, 0.234, 0.201, 0.178, 0.156, 0.134, 0.112, 0.090, 0.067, 0.045, 0.023]
    
    # Update matrix
    result = agent.update_training_matrix(1, test_cpi_array)
    
    if result['success']:
        print(f"\n✅ Matrix update completed successfully!")
        print(f"📊 Matrix dimensions: {result['matrix_dimensions']}")
        print(f"📊 Training days: {result['matrix_stats']['training_days']}")
        print(f"📊 Average CPI mean: {result['matrix_stats']['average_cpi_mean']:.6f}")
        
        # Save results
        agent.save_matrix_results(result)
        
        return True
    else:
        print(f"\n❌ Matrix update failed: {result.get('error', 'Unknown error')}")
        return False

if __name__ == "__main__":
    main()
